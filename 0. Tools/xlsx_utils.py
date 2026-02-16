"""
TheGrantScout Excel Utilities
==============================

Helper module for creating and editing Excel workbooks per CLAUDE.md standards.

Usage (creating):
    from xlsx_utils import create_workbook
    create_workbook(
        data=[["Foundation A", 500000, 0.85], ["Foundation B", 250000, 0.72]],
        headers=["Foundation", "Assets", "Score"],
        output_path="report.xlsx",
        sheet_name="Foundations",
        col_formats={"Assets": "currency", "Score": "percent"}
    )

Usage (editing):
    from xlsx_utils import edit_workbook
    edit_workbook(
        path="report.xlsx",
        edits={"B3": 600000, "C3": 0.90}
    )

Standards implemented (from CLAUDE.md):
    - Blank row 1 and column A (data starts at B2)
    - Zoom: 140%
    - Excel Table with filtering/sorting (TableStyleMedium2)
    - Numbers formatted with commas (#,##0)
    - Currency formatted as "$"#,##0
    - Percentages stored as decimal, formatted as 0%
    - Font: Calibri 10pt
    - Headers: Bold white text on dark blue (#2F5496)
    - Borders: Medium around table edges + header; thin inside
    - No wrap text
    - Freeze panes: header row + column A visible

Dependencies:
    pip3 install openpyxl
"""

import os
import sys

# Allow importing branding.py from same directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import branding

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ============================================================================
# STYLE CONSTANTS (from branding.py + CLAUDE.md Excel rules)
# ============================================================================

HEADER_FILL = PatternFill(
    start_color=branding.EXCEL_HEADER_HEX,
    end_color=branding.EXCEL_HEADER_HEX,
    fill_type="solid"
)
HEADER_FONT = Font(
    name=branding.EXCEL_FONT,
    size=branding.EXCEL_FONT_SIZE,
    bold=True,
    color="FFFFFF"
)
DATA_FONT = Font(
    name=branding.EXCEL_FONT,
    size=branding.EXCEL_FONT_SIZE,
    color=branding.CHARCOAL_BARE
)
THIN_BORDER = Side(border_style="thin", color="000000")
MEDIUM_BORDER = Side(border_style="medium", color="000000")
NO_WRAP = Alignment(wrap_text=False, vertical="center")

# Format strings
FMT_NUMBER = '#,##0'
FMT_CURRENCY = '"$"#,##0'
FMT_PERCENT = '0%'
FMT_DECIMAL = '#,##0.00'

# Format name → format string mapping
FORMAT_MAP = {
    "number": FMT_NUMBER,
    "currency": FMT_CURRENCY,
    "percent": FMT_PERCENT,
    "decimal": FMT_DECIMAL,
    "text": None,  # No special format
}


def create_workbook(data, headers, output_path, sheet_name="Sheet1", col_formats=None):
    """Create a formatted Excel workbook per TGS standards.

    Args:
        data: List of lists (rows of data, no headers).
        headers: List of column header strings.
        output_path: Path to save .xlsx file.
        sheet_name: Worksheet name.
        col_formats: Dict mapping header name → format type.
            Format types: "number", "currency", "percent", "decimal", "text"
            Example: {"Assets": "currency", "Score": "percent"}

    Returns:
        Path to created file.
    """
    if col_formats is None:
        col_formats = {}

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Zoom to 140%
    ws.sheet_view.zoomScale = branding.EXCEL_ZOOM

    start_row = branding.EXCEL_DATA_START_ROW  # 2
    start_col = branding.EXCEL_DATA_START_COL  # 2 (column B)
    num_cols = len(headers)
    num_rows = len(data)

    # Write headers at row 2, starting at column B
    for col_idx, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = NO_WRAP

    # Write data starting at row 3
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(
                row=start_row + 1 + row_idx,
                column=start_col + col_idx,
                value=value
            )
            cell.font = DATA_FONT
            cell.alignment = NO_WRAP

            # Apply column format
            header_name = headers[col_idx] if col_idx < len(headers) else None
            fmt_type = col_formats.get(header_name)
            if fmt_type and fmt_type in FORMAT_MAP and FORMAT_MAP[fmt_type]:
                cell.number_format = FORMAT_MAP[fmt_type]

    # Apply borders
    end_row = start_row + num_rows  # last data row
    end_col = start_col + num_cols - 1

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            # Determine border sides
            top = MEDIUM_BORDER if row == start_row else THIN_BORDER
            bottom = MEDIUM_BORDER if (row == start_row or row == end_row) else THIN_BORDER
            left = MEDIUM_BORDER if col == start_col else THIN_BORDER
            right = MEDIUM_BORDER if col == end_col else THIN_BORDER

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Create Excel Table for filtering/sorting
    table_start = f"{get_column_letter(start_col)}{start_row}"
    table_end = f"{get_column_letter(end_col)}{end_row}"
    table_ref = f"{table_start}:{table_end}"

    # Sanitize table name (no spaces, special chars)
    safe_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
    table = Table(displayName=safe_name, ref=table_ref)
    style = TableStyleInfo(
        name=branding.EXCEL_TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-fit column widths (approximate: max of header/data lengths + padding)
    for col_idx in range(num_cols):
        col_letter = get_column_letter(start_col + col_idx)
        max_len = len(str(headers[col_idx]))
        for row_data in data:
            if col_idx < len(row_data) and row_data[col_idx] is not None:
                cell_len = len(str(row_data[col_idx]))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Set column A width narrow (blank spacer)
    ws.column_dimensions['A'].width = 2

    # Freeze panes: keep headers and column A visible
    # Freeze at B3 = row 3, col B → headers (row 2) and col A always visible
    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)

    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def edit_workbook(path, edits, save_path=None):
    """Edit an existing workbook, preserving formatting.

    Args:
        path: Path to existing .xlsx file.
        edits: Dict mapping cell reference → new value.
            Example: {"B3": 600000, "C5": "Updated text"}
        save_path: Path to save (default: overwrite original).

    Returns:
        Path to saved file.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path)
    ws = wb.active

    for cell_ref, value in edits.items():
        ws[cell_ref] = value

    out = save_path or path
    wb.save(out)
    print(f"Updated: {out}")
    return out
