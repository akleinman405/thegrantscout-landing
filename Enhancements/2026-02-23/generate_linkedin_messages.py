#!/usr/bin/env python3
"""Generate LinkedIn connection messages for prospect spreadsheet."""

import openpyxl
import re

# Mission text → short phrase mapping (292 unique missions)
MISSION_PHRASES = {
    "TO PROVIDE JOURNEYMEN AND APPRENTICE TRAINING FOR ELECTRICAL WORKERS COVERED BY THE COLLECTIVE BARGAINING AGREEMENT BETWEEN THE UNION AND THE ELECTRICAL CONTRACTORS": "apprentice training for electrical workers",
    "INVICTUS ACADEMY OF RICHMOND PREPARES 100% OF STUDENTS IN GRADES 7-10 TO THRIVE IN THE COLLEGES OF THEIR CHOICE, SOLVE RELEVANT PROBLEMS, AND COMMUNICATE WITH CONFIDENCE.": "college prep for grades 7-10 in Richmond",
    "TO BUILD A STRONG ACADEMIC FOUNDATION THAT NURTURES OUR CHILDREN'S LIFE-LONG LOVE OF LEARNING THROUGH AN ENGAGING AND GLOBAL CURRICULUM.": "nurturing lifelong learning through a global curriculum",
    "TO SUPPORT SCHOOLS OPERATED BY THE FOUNDATION FOR HISPANIC EDUCATION AND TO ADVANCE AND DEVELOP INITIATIVES THAT BENEFIT LOW-INCOME STUDENTS ATTENDING EAST SAN JOSE SCHOOLS.": "supporting low-income students in East San Jose schools",
    "SEE SCHEDULE O": None,  # handled by org-specific overrides
    "Rivet School's mission is to build and scale a new model of higher education that reliably enables traditionally underserved students to earn an affordable bachelor's degree in as little as two years,": "affordable bachelor's degrees for underserved students",
    "The mission of The Long Now Foundation is to promote long-term thinking and to explore new ways to inspire others to think long-term. A broader understanding of the present moment can reframe enormous": "promoting long-term thinking and inspiring others to do the same",
    "THE PRIMARY PURPOSE OF THIS CORPORATION IS TO BUILD COMMUNITY AND INSPIRE POSITIVE SOCIAL CHANGE THROUGH EDUCATION ENHANCEMENT, CAREER TRAININGS, HEALTH PROMOTION, AND LEADERSHIP DEVELOPMENT WITH LESB": "youth leadership development and community building",
    "THE LITTLE SCHOOL IS A PRIVATE NON-PROFIT PRESCHOOL FOUNDED TO PROVIDE YOUNG CHILDREN IN SAN FRANCISCO WITH A DEVELOPMENTAL, CHILD-CENTERED PRESCHOOL EXPERIENCE. THE LITTLE SCHOOL IS COMMITTED TO PROV": "child-centered preschool education in San Francisco",
    "Transformations of the human is a platform for DOING philosophy. We provide philosophical insight, aesthetically rendered, in an effort to build products practitioners and companies with philosophical": "philosophy-driven education and leadership",
    "Education: Preschool,Elementary (K-5), Daycare": "preschool and elementary education",
    "TO EDUCATE APPRENTICES AND JOURNEYMEN IN THE ELECTRICAL INDUSTRY WHO ARE COVERED BY A COLLECTIVE BARGAINING AGREEMENT.": "apprentice education in the electrical industry",
    "THE AVALON ACADEMY IS A SCHOOL DEDICATED TO PROVIDING EXCEPTIONAL EDUCATIONAL SERVICES TO CHILDREN WITH MOVEMENT DISORDERS SUCH AS CEREBRAL PALSY.": "education for children with movement disorders like cerebral palsy",
    "NONE": None,  # handled by org-specific overrides
    "THE FOUNDATION FOR SFJAZZ IS A CA NON-FOR-PROFIT CORPORATION ORGANIZED ON SEPTEMBER 14, 2011. THE FOUNDATION, WHOSE SOLE MEMBER IS SAN FRANCISCO JAZZ ORGANIZATION, WAS SPECIFICALLY ORGANIZED FOR THE E": "jazz education and performance in San Francisco",
    "The School of Imagination continues to lead the way in inclusive early childhood education and multidisciplinary intervention programs for neurodiverse learners. Our mission remains steadfast: to brin": "inclusive early childhood education for neurodiverse learners",
    "Coleman builds the leadership and power of low-income and working-class BIPOC children, youth, and families in San Francisco to advance racial and economic justice in our schools and our city.": "building leadership and power in BIPOC youth and families",
    "MONTESSORI DE TERRA LINDA IS A DIVERSE COMMUNITY OF TEACHERS, PARENTS AND CHILDREN WORKING TOGETHER FOR THE EDUCATION OF THE CHILD, UTILIZING THE METHODS OF MARIA MONTESSORI. WE SERVE CHILDREN AGES 2-": "Montessori education for children ages 2 and up",
    "NEW BREATH FOUNDATION PROVIDES PIVOTAL GRASSROOTS ORGANIZATIONS THE SUSTAINABLE RESOURCES THEY NEED TO GIVE NEW BEGINNINGS TO AANHPIS HARMED BY INCARCERATION, DEPORTATION, AND SYSTEMS OF VIOLENCE IN T": "supporting AANHPI communities impacted by incarceration",
    "Comprehensive family service system.": "comprehensive family services",
    "FAMILY SUPPORT SERVICES'S PROGRAMS SERVE FAMILIES, YOUTH AND CHILDREN WHOSE HEALTH AND WELFARE ARE VULNERABLE AS A RESULT OF A VARIETY OF CIRCUMSTANCES. WE PROVIDE ASSISTANCE TO FAMILIES STRUGGLING TO": "supporting vulnerable families, youth, and children",
    "THE ACADEMY'S PRIMARY PURPOSE IS TO EDUCATE CHILDREN IN KINDERGARTEN THROUGH EIGHTH GRADE BY DEVELOPING THOUGHTFUL, AMBITIOUS, AND DEEP LEARNERS. THE ACADEMY HONORS EACH STUDENT'S INDIVIDUALITY WHILE": "developing thoughtful, deep learners in grades K-8",
    "TO INSPIRE COMMUNITIES AND ENRICH LIVES BY INCREASING OPPORTUNITIES FOR PARTICIPATION, EDUCATION, AND EXCELLENCE IN THE PERFORMING ARTS.": "performing arts education and community engagement",
    "SERVING YOUNG CHILDREN WITH SPECIAL NEEDS AND THEIR FAMILIES.": "serving young children with special needs and their families",
    "IMPROVING EDUCATION IN DEVELOPING COUNTRIES THROUGH INNOVATIVE TECHNOLOGY SOLUTIONS. ACTING AS A CATALYST FOR CHANGE BY MATCHING TECH INDUSTRY PROFESSIONALS, PRODUCTS AND RESOURCES WITH ACCREDITED NON": "improving education in developing countries through technology",
    "THE PHILLIPS ACADEMY'S MISSION IS TO PROVIDE A PERSONALIZED EDUCATIONAL EXPERIENCE FOR STUDENTS WITH DIVERSE AND COMPLEX LEARNING AND EMOTIONAL STYLES, BY ADDRESSING THE NEEDS OF THE WHOLE CHILD. UPON": "personalized education for students with diverse learning styles",
    "TO PREPARE CHILDREN ON THE AUTISM SPECTRUM FOR A FULFILLING FUTURE THROUGH DYNAMIC RELATIONSHIP-BASED INTERDISCIPLINARY EDUCATION.": "relationship-based education for children on the autism spectrum",
    "SMARTMEME INC. DBA CENTER FOR STORY-BASED STRATEGY (\"CSS\") IS A NATIONAL STRATEGY CENTER DEDICATED TO HARNESSING THE POWER OF NARRATIVE FOR MOVEMENT BUILDING AND SOCIAL CHANGE. CSS PROVIDES SOCIAL, EC": "story-based strategy for social change movements",
    "The Life Learning Academy is committed to creating a nonviolent community for students who have not been successful in traditional school settings. Life Learning Academy welcome students into an 'exte": "creating community for students outside traditional school settings",
    "CLOSE THE DIGITAL DIVIDE FOR THE 18 MILLION HOUSEHOLDS THAT HAVE ACCESS TO THE INTERNET BUT CAN'T AFFORD TO CONNECT. WE FOCUS ON AMERICA'S MOST UNCONNECTED COMMUNITIES, WHERE MORE THAN 25% OF PEOPLE D": "closing the digital divide for unconnected communities",
    "Empower underserved students and their teachers through mindfulness and other transformative skills to gain self-awareness, confidence, self-regulation and resilience, leading to lifelong success.": "mindfulness education for underserved students and teachers",
    "PROVISION OF CHILD CARE SERVICES.": "child care services for families",
    "SCHOLARMATCH'S MISSION IS TO SUPPORT UNDERSERVED FIRST-GENERATION COLLEGE STUDENTS FROM LOW-INCOME BACKGROUNDS TO EARN A BACHELOR'S DEGREE WITHIN FIVE YEARS.": "supporting first-generation college students from low-income backgrounds",
    "To challenge students to think creatively, act compassionately, and live courageously.": "cultivating creative and compassionate thinkers",
    "Work with community partners to provide & facilitate apprenticeship programs educate the public about high-wage construction careers": "construction apprenticeship programs and career education",
    "THE MISSION OF SAN MATEO COUNTY COMMUNITY COLLEGES FOUNDATION IS TO PROMOTE STUDENT SUCCESS AND PROGRAM INNOVATION BY PROVIDING SPECIAL FINANCIAL SUPPORT TO HELP SAN MATEO COUNTY COMMUNITY COLLEGE DIS": "student success and scholarships at San Mateo community colleges",
    "Our mission is to build character and leadership skills through a combination of training, practice, and service. Students learn greater world awareness as well as how leadership can enhance their lea": "building character and leadership skills through service",
    "Established in 2011, our mission is to help transform the states higher education system into an engine of economic opportunity that empowers all Californians, particularly those from underserved comm": "transforming higher education access for all Californians",
    "COW HOLLOW SCHOOL IS A PLAY BASED PART-DAY PROGRAM, SERVING CHILDREN AGED 2-5.9 YEARS OLD. WE ARE A PARENT PARTICIPATION SCHOOL WITH A LOW TEACHER TO CHILD RATIO AND LICENSED MASTER TEACHERS ON STAFF.": "play-based preschool with low teacher-to-child ratios",
    "Wellspring Educational Services was founded to provide support to families who desire a developmentally integrative approach to educating their child.": "developmentally integrative education for families",
    "Kingmakers of Oakland is a leadership development and systems change organization committed to fundamentally transforming the education system and building the capacity of people to design and sustain": "transforming education systems and building youth leadership in Oakland",
    "Empower, educate, inspire, and enlighten tomorrows generation.": "empowering and educating tomorrow's generation",
    "The mission of All Five is to empower all families to choose a high quality early childhood education for their children. We offer a full-day, year-around program for children and their families betwe": "high-quality early childhood education for all families",
    "TO SUPPORT QUALITY EDUCATIONAL PROGRAMS IN THE ARTS, DANCE, DRAMA, MUSIC AND IN PHYSICAL EDUCATION FOR EACH CHILD IN THE MILL VALLEY SCHOOL DISTRICT.": "arts, dance, drama, and music programs for Mill Valley students",
    "Challenge Success partners with schools, families, and communities to embrace a broad definition of success and implement research-based strategies that promote student well-being and engagement with": "research-based strategies for student well-being",
    "THE ORGANIZATION'S MISSION IS TO PROVIDE AND MANAGE THE CAMPUS BOOKSTORE, DINING SERVICES AND CERTAIN REAL ESTATE PROPERTIES. SPARTAN SHOPS HAS AN AGREEMENT WITH BARNES & NOBLE COLLEGE BOOKSELLERS (\"B": "campus services supporting SJSU students",
    "CONNECTED: THE NATIONAL CENTER FOR COLLEGE AND CAREER PARTNERS WITH SCHOOL, DISTRICT, AND COMMUNITY LEADERS TO TRANSFORM EDUCATION THROUGH LINKED LEARNING PATHWAYS SO THAT ALL STUDENTS, REGARDLESS OF": "college and career pathways through Linked Learning",
    "TO PROVIDE AN AUTHENTIC MONTESSORI EDUCATION IN A NURTURING SCHOOL COMMUNITY. WITH A STRONG EMPHASIS ON ACADEMICS, CREATIVITY AND PEACE EDUCATION, WE PROVIDE A LEARNING EXPERIENCE THAT PROMOTES CONFID": "Montessori education with creativity and peace education",
    "Oakland Leaf spurs creative and educational growth for oakland youth by providing programs that cultivate critical thinking skills, emphasize socioemotional learning opportunities, and support authent": "creative and educational programs for Oakland youth",
    "TO ENGAGE AND SERVE STANFORD GSB ALUMNI THROUGHOUT THEIR LIVES AND TO ADVANCE STANFORD GSB IN ITS MISSION TO DEVELOP INNOVATIVE, PRINCIPLED, AND INSIGHTFUL LEADERS WHO CHANGE THE WORLD.": "developing innovative leaders through Stanford GSB",
    "TO SUPPORT YOUNG PEOPLE AS THEY IMAGINE AND CREATE THEIR OWN LIFE BEYOND FOSTER CARE.": "supporting young people transitioning out of foster care",
    "CDE Foundation works as a trusted partner with state education leaders and entities to create, resource, and implement solutions that result in a strong and valued public education system that serves": "strengthening California's public education system",
    "TO PROVIDE A PROGRAM FOR THE EDUCATION OF APPRENTICES AND SUPPLEMENT EDUCATION OF JOURNEYMEN IN THE ELECTRICAL INDUSTRY.": "apprentice and journeyman education in the electrical industry",
    "Since 1978, ECC has grown from a small school age program on two school campuses, to one of the largest after school programs in Sonoma County. Quality programming has been the agency's guiding force": "after-school programs across Sonoma County",
    "We strive to improve cancer health outcomes, FDA cancer clinical trial diversity and enrollment, and patient access to care by providing assistance with clinical trial navigation, reimbursing trial re": "improving access to cancer clinical trials",
    "LLS contracts with schools to provide after school and summer child care programs for elementary school children. LLS's programs are designed to encourage safe, nurturing and educationally enriching a": "after-school and summer programs for elementary students",
    "THE EDUCATION FUND MOBILIZES THE COMMUNITY TO SUPPORT EQUITABLE ACCESS TO A QUALITY EDUCATION FOR PUBLIC SCHOOL STUDENTS THROUGH TUTORING AND MENTORING, SCHOLARSHIPS, TEACHER GRANTS, AND CORPORATE SCH": "tutoring, mentoring, and scholarships for SF public school students",
    "TO PROVIDE CHILD CARE AND FAMILY SERVICES TO LOW AND MODERATE INCOME FAMILIES WHO LIVE OR WORK IN CONCORD, CA.": "child care for low-income families in Concord",
    "CRISTO REY SAN JOSE WORK STUDY CORPORATION SECURES AND MANAGES PROFESSIONAL WORK EXPERIENCES FOR HIGH SCHOOL STUDENTS AT CRISTO REY SAN JOSE HIGH SCHOOL. STUDENTS WORK FOR 100+ SILICON VALLEY COMPANIE": "professional work experiences for high school students in Silicon Valley",
    "TO HELP SUPPLEMENT THE FUNDS OF THE BELMONT-REDWOOD SHORES SCHOOL DISTRICT SO THAT PROGRAMS CAN CONTINUE TO BE RETAINED. THESE PROGRAMS MAY INCLUDE SUCH DISTRICT-WIDE ENRICHMENT PROGRAMS AS: MUSIC PRO": "music and enrichment programs in Belmont-Redwood Shores schools",
    "SEE SCHEDULE OSHADOWSERVER'S MISSION IS TO MAKE THE INTERNET MORE SECURE BY BRINGING TO LIGHT VULNERABILITIES, MALICIOUS ACTIVITY AND EMERGING THREATS. SINCE 2004, SHADOWSERVER HAS BEEN INVESTIGATING": "internet security research and threat detection",
    "Provide an alternative public education experience for children in the Walnut Creek, CA area.": "alternative public Montessori education in Walnut Creek",
    "Family Builders educates the community about the needs of waiting children, advocates on their behalf, and places children with permanent, secure families through adoption, and other forms of permanen": "placing children in permanent families through adoption",
    "Our mission is to spark wonder and intellectual curiosity in each child. We achieve this through a purposeful blend of:Engaging academics custom-tailored to each child's abilities, strengths and inter": "sparking wonder and intellectual curiosity in each child",
    "The mission of The Edible Schoolyard Project is to build and share a national food curriculum for the education system. The Organization envisions this \"Edible Education\" as part of the core curriculu": "hands-on food education and school garden curriculum",
    "The mission of the organization is to provide public school students of southern Sonoma County with a K-8 whole- child Public Waldorf educational program. The organization embraces a developmental app": "whole-child Waldorf education in southern Sonoma County",
    "TO EDUCATE HIGH FUNCTIONING DYSLEXIC CHILDREN.": "educating children with dyslexia",
    "TO SUPPORT THE ACROSS THE BRIDGE FOUNDATION DBA DOWNTOWN COLLEE PREPARATORY (ABF) AND THE CHARTER SCHOOLS IT OPERATES BY PROVIDING REAL ESTATE DEVELOPMENT AND MANAGEMENT SERVICES INCLUDING PLANNING, F": "supporting college prep charter schools in downtown San Jose",
    "THE NOVATO CHARTER SCHOOL IS AN EDUCATIONAL COMMUNITY THAT IS COMMITTED TO THE HEALTHY GROWTH AND DEVELOPMENT OF THE WHOLE CHILD. THROUGH A CURRICULUM INSPIRED BY WALDORF-METHODS, OUR TEACHERS NURTURE": "Waldorf-inspired whole-child education in Novato",
    "TO EDUCATE STUDENTS WITH LEARNING DIFFERENCES FOCUSING ON THEIR INDIVIDUAL NEEDS WITH THE GOAL OF MAXIMIZING THEIR INHERENT ABILITIES AND THEIR POTENTIAL TO ACHIEVE SUCCESS. BAYHILL SCHOOL HAS AN INDI": "individualized learning for students with learning differences",
    "The Oakland REACH is a parent-run, parent-led group committed to empowering families from our most underserved communities to demand high-quality schools for our children.": "empowering Oakland families to demand high-quality schools",
    "Provide low cost Christian education": "affordable Christian education",
    "See Schedule O.": None,  # handled by org-specific overrides
    "TO PROVIDE AFFORDABLE HOUSING TO STUDENTS": "affordable student housing",
    "The mission of Bridge the Gap is to provide comprehensive educational, social and emotional resources to underserved students in Marin City and Southern Marin. Our goal is to ensure that all students": "educational and social-emotional support for students in Marin City",
    "ROCK'S MISSION IS TO NURTURE THE HEALTHY DEVELOPMENT OF VISITACION VALLEY YOUTH BY LISTENING TO THEIR NEEDS AND BY PROVIDING OPPORTUNITIES TO THOSE WHO MIGHT NOT OTHERWISE HAVE ACCESS. THIS IS ACCOMPL": "nurturing healthy development in Visitacion Valley youth",
    "Math Science Nucleus is a non-profit 501c3 educational and research organization composed of scientists educators and community members founded in 1982. It serves as an online science resource center": "hands-on science education and community research",
    "IMPACTING GENERATIONAL CHANGE BY EMPOWERING YOUTH WHO ARE FACING THE GREATEST OBSTACLES THROUGH RELATIONSHIPS WITH PROFESSIONAL MENTORS - 12+ YEARS, NO MATTER WHAT.": "12-year professional mentoring for youth facing obstacles",
    "PCF ASSISTS LOW-INCOME, UNDER SERVED AND OFTEN OVER-LOOKED FIRST-GENERATION COLLEGE STUDENTS FROM THE MID-PENINSULA IN THEIR QUEST TO GRADUATE, AND ASSISTS THEM TO ACHIEVE THEIR EDUCATIONAL AND CAREER": "first-generation college student support on the Peninsula",
    "We believe families make San Francisco stronger and more vibrant. We are committed to helping create a stronger, more vibrant San Francisco by ensuring that Potrero Hill, Dogpatch and Mission Bay fami": "supporting families in Potrero Hill, Dogpatch, and Mission Bay",
    "OPERATES AN INDEPENDENT MIDDLE SCHOOL FOR BOYS (GRADES 6-8).": "independent middle school for boys in grades 6-8",
    "THROUGH RIGOROUS INSTRUCTION AND POSITIVE CHARACTER EDUCATION, AURUM PREPARATORY ACADEMY CHARTER SCHOOL EDUCATES ALL OF ITS STUDENTS, GRADES 6 -8, TO SUCCEED IN HIGH SCHOOL, COLLEGE, AND LIFE AND TO S": "rigorous academics and character education for grades 6-8",
    "BURLINGAME COMMUNITY FOR EDUCATION FOUNDATION (BCE) HELPS BURLINGAME SCHOOL DISTRICT (BSD) SUSTAIN AND ENHANCE AN EXCEPTIONAL PUBLIC EDUCATION FOR ALL TK-8TH GRADE STUDENTS. WORKING TOGETHER WITH PARE": "enhancing public education for Burlingame students",
    "TO PREPARE STUDENTS TO BE SELF-AWARE, RESPONSIBLE, LIFE-LONG LEARNERS, MOTIVATED TO MEET THE CHALLENGES OF A RAPIDLY CHANGING WORLD. TO DEVELOP FUTURE LEADERS BY EXPECTING ACADEMIC EXCELLENCE, ENHANCI": "preparing self-aware, responsible lifelong learners",
    "MG advances a Just Transition away from profit and pollution, towards healthy and resilient local economies. We've engaged over 160 organizations and thousands of change agents through retreats, polit": "advancing a just transition to resilient local economies",
    "TOIGO DEVELOPS LEADERS AND FOSTERS THE CAREER ADVANCEMENT OF HIGHLY TALENTED WOMEN AND MEN THROUGHOUT THEIR MBA EXPERIENCE. OUR NEARLY 2,000 TOIGO ALUMNI ARE THE FOUNDATION ON WHICH WE HAVE BUILT PROG": "career advancement for talented women and diverse MBA leaders",
    "To educate children ages 18 months to 12 years using Montessori principles and other innovative teaching practices.": "Montessori education for ages 18 months to 12 years",
    "TO SERVE LOW INCOME FAMILIES WITH CHILDREN: SUPPORTING SELF SUFFICIENCY, FOSTERING HEALTH AND ACTIVE LIFESTYLES, AND INSPIRING JOY IN LEARNING.": "supporting low-income families and inspiring joy in learning",
    "ALL CHILDREN SHUOLD HAVE ACCESS TO A QUALITY EDUCATION THAT PROVIDES A BROAD RANGE OF EXPERIENCES.STUDENTS SHOULD ACTIVELY PARTICIPATE IN THEIR OWN LEARNING.TEACHING SHOULD BE BASED ON THE LEARNING ST": "active, experience-driven education for all children",
    "THE ORGANIZATION PROVIDES EARLY-CHILDHOOD EDUCATION WITH AN ART BASED PROGRAM THAT PROMOTES INDIVIDUAL GROWTH, SELF-AWARENESS, AND READINESS FOR FORMAL LEARNING.": "art-based early childhood education",
    "THE OXBOW SCHOOL IS A UNIQUE, INTERDISCIPLINARY SEMESTER PROGRAM FOR HIGH SCHOOL STUDENTS. OUR MISSION IS TO STRENGTHEN STUDENTS' ABILITIES IN CREATIVE AND CRITICAL INQUIRY BY COMBINING RIGOROUS STUDI": "interdisciplinary semester programs for high school students",
    "The mission of Mount Tamalpais College is to provide an intellectually rigorous, inclusive Associate of Arts degree program and College Preparatory Program, free of charge, to people at San Quentin St": "free college education for incarcerated students at San Quentin",
    "To enrich San Francisco and the world by providing best practice childcare servicesto culturally and economically diverse families while serving parents through anenduring community of staff, peer par": "best-practice childcare for diverse families in San Francisco",
    "RAFT BELIEVES THAT EVERY CHILD DESERVES A POWERFUL LEARNING EXPERIENCE. RAFT'S MISSION IS TO INSPIRE JOY THROUGH HANDS ON LEARNING.": "inspiring joy through hands-on learning",
    "Were working to rapidly reduce climate pollution at scale, starting in California.": "reducing climate pollution at scale in California",
    "Our mission is to prepare children for Kindergarten and beyond by promoting educational achievement, ensuring household economic stability and strengthening cultural identity. We fulfill our mission b": "kindergarten readiness and cultural identity in Chinatown",
    "THE CENTER'S MISSION IS TO SUPPORT WORKING FAMILIES BY PROVIDING HIGH QUALITY CHILD CARE FOR CHILDREN AGED SIX WEEKS TO SIX YEARS.": "high-quality child care for ages six weeks to six years",
    "Educational programs include Toddler, AM Primary, Full Day, Kindergarten, Summer School, Afterschool, and Extended Day Care. Approximately 115 families.": "Montessori programs from toddler through kindergarten",
    "HOUSING REHABILITATION AND REPAIRS": "housing rehabilitation and repairs for families",
    "THE MISSION OF OPEN PHILANTHROPY IS TO STUDY AND TO ENGAGE IN RESEARCH ANALYZING THE MOST PROMISING OPTIONS FOR ENSURING THAT PHILANTHROPY HAS A LASTING IMPACT ON IMPROVING THE LIVES OF AS MANY PEOPLE": "high-impact philanthropy research",
    "AI4ALL'S MISSION IS TO ENSURE THAT THE NEXT GENERATION OF AI LEADERS REFLECTS HUMANITY. AI4ALL IS TRANSFORMING THE PIPELINE OF AI PRACTITIONERS WHO WILL SHAPE AI FOR THE BENEFIT OF HUMANITY.": "building the next generation of diverse AI leaders",
    "Step One School was established in 1981 in Berkeley, California to provide the East Bay community with a model early childhood education program for children ages two to six. We are committed to givin": "model early childhood education in the East Bay",
    "TANDEM, PARTNERS IN EARLY LEARNING IS A BAY AREA NONPROFIT WORKING AT THE INTERSECTION OF SOCIAL JUSTICE AND EARLY CHILDHOOD EDUCATION. WE WORK ALONGSIDE SCHOOL DISTRICTS, EARLY CHILDHOOD EDUCATION AG": "social justice and early childhood education",
    "A LEADER IN THE BAY AREA ARTS COMMUNITY, THE CRUCIBLE IS WIDELY RECOGNIZED FOR ITS EXCEPTIONAL LEARNING EXPERIENCES, RICH AND VARIED ARTS PROGRAMS, SKILLED AND COMMITTED FACULTY, AND UNPARALLELED EDUC": "hands-on industrial arts education in the Bay Area",
    "To advocate for and advance the accessibility of quality education for all through sustainable school models.": "accessible quality education through sustainable school models",
    "THE CHARTER SCHOOL WILL EMPOWER EVERY STUDENT TO BE A LIFELONG LEARNER WITH ESSENTIAL SKILLS AND KNOWLEDGE UTILIZING AN EDUCATIONAL PROGRAM INSPIRED BY WALDORF EDUCATION METHODS.": "Waldorf-inspired education for lifelong learners",
    "IN COLLABORATION WITH COMMUNITY PARTNERS, OUR MISSION IS TO SUSTAIN THE LAFAYETTE LIBRARY AND LEARNING CENTER, HOME OF THE GLENN SEABORG LEARNING CONSORTIUM, AS A COMMUNITY PLACE AND REGIONAL RESOURCE": "community learning programs at the Lafayette Library",
    "THE MISSION OF BLUESKIES IS TO PRACTICE AND TEACH AN OPTIMAL APPROACH TO EDUCATION AND CARE OF YOUNG CHILDREN BASED ON THE INTERRELATED DISCIPLINES OF HUMAN DEVELOPMENT. SUCH AN APPROACH SUSTAINS AND": "research-based early childhood education and care",
    "TRANSFORMING THE LIVES OF INDIVIDUALS WITH MODERATE TO SEVERE AUTISM AND THEIR FAMILIES THROUGH A CUSTOMIZED APPROACH TO LIFETIME SERVICES, EDUCATIONAL PROGRAMS AND ADVOCACY.": "lifetime services and education for individuals with autism",
    "THE NONPROFIT ASPIRE PUBLIC SCHOOLS FOUNDATION IS THE PHILANTHROPIC ARM OF ASPIRE PUBLIC SCHOOLS. THE FOUNDATION EXISTS TO PROVIDE FINANCIAL AND MATERIAL RESOURCES TO FURTHER ASPIRE'S MISSION OF PREPA": "preparing students in under-resourced communities for college",
    "THE SCHOOL OFFERS TO ITS STUDENTS A COMPLETE, LOVING AND DEVELOPMENTAL APPROACH TO EDUCATION; AN EDUCATION THAT HONORS THE INDIVIDUAL AS MUCH AS IT TEACHES THE IMPORTANCE OF COMMUNITY.": "developmental, community-centered education",
    "MEF's goal is to support and enrich the educational experience of school children in kindergarten through twelfth grade.": "enriching K-12 education in Moraga",
    "Athena Academy Inc.'s (Academy) mission is dedicated to developing and providing the most effective methods available to teach dyslexic children, so they may develop their gifts to learn and excel. Th": "specialized teaching methods for dyslexic children",
    "TO ENSURE THE SUCCESSFUL DEVELOPMENT AND FINANCIAL MANAGEMENT OF THE CAMPUS FACILITIES.": "campus development and management",
    "THE EXEMPT PURPOSE OF THIS CORPORATION IS TO FURTHER THE EDUCATION AND TRAINING OF YOUNG PEOPLE IN A MANNER CONSISTENT WITH A CERTIFIED CURRICULUM.": "youth education and training programs",
    "Nihonmachi Little Friends (NLF) is a Japanese bilingual multicultural & educational childcare organization.": "Japanese bilingual multicultural childcare",
    "GEOKIDS IS AN ACCREDITED, NATIONALLY RECOGNIZED, NON-PROFIT PARENT COOPERATIVE PROVIDING HIGH-QUALITY CHILD CARE FOR CHILDREN AGED 3 MONTHS THROUGH PRE-KINDERGARTEN. GEOKIDS IS A DEVELOPMENTAL PLAY-BA": "developmental play-based childcare for infants through pre-K",
    "TRANSFORM LIVES OF CHILDREN WHO ARE DEAF OR HARD OF HEARING, TEACHING THEM TO LISTEN, SPEAK, AND LEARN THROUGH PROVEN SPOKEN LANGUAGE AND COGNITIVE DEVELOPMENT CURRICULUM, FAMILY-CENTERED EDUCATION AN": "teaching deaf and hard-of-hearing children to listen and speak",
    "To champion equitable public education for students, families, and our community. Through grants, volunteers, and STEM programs, we center equity for the benefit of our students and the entire communi": "equitable public education and STEM programs in Berkeley",
    "PROVIDE NBSP;HIGH NBSP;QUALITY NBSP;ON-SITE NBSP;CHILD NBSP;CARE NBSP;FOR NBSP;THE NBSP;INFANTS NBSP;TODDLERS NBSP;AND NBSP;PRESCHOOLERS NBSP;OF NBSP;THE NBSP;CIVIL NBSP;SERVANTS NBSP;AND NBSP;CONTRAC": "on-site child care for NASA Ames families",
    "The Oakland Museum Women's Board is a California non-profit, public benefit corporation whose sole purpose is to provide services and funds to the Oakland Museum of California.": "supporting the Oakland Museum of California",
    "Partner and invest in the people, organizations, and projects working to build civic gathering spaces around the country.": "building civic gathering spaces across the country",
    "ISKME's mission is to make learning and knowledge-sharing participatory, equitable and open.": "participatory, equitable, and open learning",
    "To provide educational and supportive services for young children, including by training and cultivating a sustainable and diverse pool of substitute teachers that can better take care of young childr": "training early childhood substitute teachers",
    "TO COMBAT SYSTEMIC INEQUITY IN PUBLIC EDUCATION BY FOSTERING AND CULTIVATING THE ACADEMIC AND SOCIAL EMOTIONAL SKILLS OF OUR SCHOLARS WHILE EXPOSING THEM TO ENGAGING PROJECT-BASED LEARNING AND COMMUNI": "combating education inequity through project-based learning",
    "The mission of Southeast Asian Development Center (SEADC) is to foster a healthy, thriving and self-sufficient Southeast Asian American community.": "fostering a thriving Southeast Asian American community",
    "PROVIDE CULTURAL AND EDUCATIONAL SERVICES TO THE PUBLIC BY PROVIDING GRADES 6-12 EDUCATION WITH BOTH ACADEMIC AND ARTISTIC TRAINING.": "grades 6-12 education combining academics and artistic training",
    "Playgroup provides a full-day Montessori pedagogy program for preschoolers, 5 days a week from 8 am-6 pm. The program blends academics, play, gross motor skills, creative arts, and music to prepare ch": "full-day Montessori preschool blending academics and play",
    "To provide a safe, inclusive community that encourages bright young people to be themselves while building character, intellect, and creativity.": "building character, intellect, and creativity in young people",
    "RESIDENTIAL CARE AND TREATMENT OF EMOTIONALLY AND PHYSICALLY NEGLECTED CHILDREN.": "residential care for emotionally and physically neglected children",
    "THE ORGANIZATION IS DEDICATED TO SERVING AND REACHING OUT TO ALL GANG IMPACTED AS WELL AS AT-RISK YOUTH, YOUNG ADULTS, THEIR FAMILIES AND COMMUNITIES, BY MEANS OF EDUCATIONAL PROGRAMS, INDIVIDUAL AND": "educational programs for gang-impacted and at-risk youth",
    "ENTERPRISE FOR YOUTH EMPOWERS UNDER-RESOURCED SAN FRANCISCO YOUTHS TO REACH THEIR POTENTIAL THROUGH TRANSFORMATIVE PAID INTERNSHIP EXPERIENCES SUPPORTED BY A COMMUNITY OF EMPLOYERS, CARING ADULTS, AND": "paid internships for under-resourced San Francisco youth",
    "An intentionally small community where young people are valued as individuals. A compassionate team of educators approach academic & social-emotional development with the aim of cultivating justice-se": "justice-centered education in a small school community",
    "THE SCHOOL WAS ORGANIZED TO PROVIDE A PRE-SCHOOL EDUCATIONAL EXPERIENCE AND OPPORTUNITIES FOR YOUNG CHILDREN IN THE GREATER SAN FRANCISCO BAY AREA.": "preschool education in the San Francisco Bay Area",
    "(continued from page 1) nationally and internationally. Further, J seeks to enrich the cultural, religious and social life of the community through its articles, interviews and features.": "enriching Jewish community life in San Francisco",
    "THE AGENCY WAS ESTABLISHED TO PROVIDE CHILDCARE EDUCATIN AND SUPPORT SERVICES FOR CHILDREN AND FAMILIES. THE AGENCY FORCUSES ON INFANT AND CHILD DEVELOPMENT AS WELL AS FAMILY DYNAMICS AND PARENT SERVI": "childcare, infant development, and family support services",
    "Golesetan conducts Persian language and Persian cultural immersion classes, exposing children to language movement, science, math, nature, theater, art, music, cooking, dance, gardening, yoga, and mor": "Persian language and cultural immersion for children",
    "THE ORGANIZATION IS DEDICATD TO PROVIDING PRE-SCHOOL WITH A THEME BASED CURRICULUM OVER A TRADITIONAL SCHOOL YEAR FOR CHILDREN AGES TWO YEARS SEVEN MONTHS THROUGH FIVE YEARS SIX MONTHS.": "theme-based preschool curriculum for young children",
    "THE MISSION OF JOINT VENTURE IS TO FOSTER THE HEALTH AND VITALITY OF SILICON VALLEY FOR ALL THE PEOPLE WHO LIVE AND WORK HERE. WE PURSUE OUR MISSION BY (1) PROVIDING DATA, RESEARCH AND ANALYSIS THROUG": "fostering the health and vitality of Silicon Valley",
    "OUR MISSION IS TO ENGAGE SYRIANS LIVING ABROAD IN PROJECTS THAT WILL UNLOCK THE POTENTIAL OF SYRIA'S ECONOMY AND ENHANCE THE LIVELIHOOD OF SYRIA'S CITIZENS. WE ARE ONE OF THE LARGEST, MOST TALENTED AN": "unlocking potential for Syrian communities through education",
    "The world is grappling with complex problems, including climate change, war, poverty, and ongoing human rights violations. These issues disproportionately impact women, particularly in developing coun": "educating women leaders from post-conflict countries",
    "Our mission is to transform lives through teaching digital literacy and equity. The ability to use digital tools to find, analyze, create and communicate information is a critical skill for the surviv": "digital literacy and equity education",
    "THE LINKED LEARNING ALLIANCE IS THE ENGINE THAT DRIVES THE LINKED LEARNING MOVEMENT, A FIELD OF EDUCATORS, EMPLOYERS, STUDENTS, AND COMMUNITY LEADERS THAT SUPPORTS EDUCATION SYSTEMS AS THEY ENGAGE YOU": "career-connected learning through the Linked Learning movement",
    "TO BRING ABOUT A TRANSFORMATION IN THE LIVES OF BRILLIANT BUT FINANCIALLY NEEDY STUDENTS IN INDIA.THE MISSION IS ACCOMPLISHED BY PROVIDING FINANCIAL ASSISTANCE TO ENABLE STUDENTS TO CONTINUE AND COMPL": "scholarships for brilliant students in India",
    "A PUBLIC SCHOOL DEDICATED TO HELPING CHILDREN TO BECOME THOUGHTFUL, INFORMED, AND INQUISITIVE CITIZENS. NOCCS IS A VIBRANT, DIVERSE LEARNING COMMUNITY DRIVEN BY RESPECT FOR EACH CHILD'S UNIQUE INTELLI": "nurturing thoughtful, inquisitive learners in North Oakland",
    "THE ORGANIZATION CARRIES OUT ITS PRIMARY MISSION THROUGH THE OPERATION OF CELEBRATION CENTER, WHICH CONSISTS OF A PRESCHOOL, AFTERSCHOOL PROGRAM, GRADES KINDERGARTEN THROUGH FOURTH GRADE. THE CENTER O": "preschool through 4th grade education and afterschool programs",
    "The Child Care Law Center (\"CCLC\") educates, advocates, and litigates to make child care a civil right.": "advocating to make child care a civil right",
    "The specific purpose of the Corporation shall include, without limitation, providing academic and cultural activities to school-age children during after-school hours and school breaks.": "after-school academic and cultural activities for children",
    "PLACEMENT OF CHILDREN AGES 0-18 WITH SAFE HEALTHY FOSTER FAMILIES HOMES": "placing children in safe, healthy foster homes",
    "OUR MISSION AS A GLOBAL ALLIANCE IS TO PROVIDE STRATEGIES, RESEARCH, AND RESOURCES THAT SUPPORT THE CRITICAL AGENCY AND INFLUENCE OF WOMEN'S FOUNDATIONS AND GENDER JUSTICE FUNDERS IN THE MOVEMENT FOR": "women's foundations and gender justice funding",
    "Silicon Valley Academy's mission is to empower students to meet their highest potential, nurture their social-emotional growth, and foster their commitment to serve society. As a full-time pre-kinderg": "empowering students from pre-K through middle school",
    "THE SPECIFIC PURPOSE OF AGM IS TO PROVIDE GROUP HOME RESIDENTIAL CARE IN VARIOUS LOCATIONS FOR CHILDREN WITH SOCIAL, EMOTIONAL, MENTAL HEALTH ISSUES, AND TO PROVIDE DRUG AND ALCOHOL RECOVERY SERVICES.": "residential care for children with mental health challenges",
    "THE MISSION OF THE FOUNDATION IS TO RAISE FUNDS FROM CURRENT AND ALUMNI FAMILIES, COMMUNITY MEMBERS, AND LOCAL FOUNDATIONS INTERESTED IN SUPPORTING EDUCATIONAL EXCELLENCE IN THEIR COMMUNITY. RATHER TH": "supporting educational excellence at Menlo-Atherton",
    "826 National amplifies the impact of our national network of youth writing and publishing centers, and the words of young authors. We serve as an international proof of point for writing as a tool for": "youth writing and publishing centers nationwide",
    "THE ORGANIZATION ACHIEVES ITS PURPOSE TO ADVANCE EDUCATION BY ENABLING THE EDUCATION AND DEVELOPMENT OF YOUTH, MINORITIES, VETERANS AND OTHER UNDERREPRESENTED POPULATIONS INTO MEANINGFUL STEM-BASED ED": "STEM education for youth, minorities, and veterans",
    "Old Skool Cafe is a faith-based violence prevention program, providing marketable and transferable employment skills to underserved youth. Our social enterprise restaurant is a hub for youth employmen": "youth employment and violence prevention through a social enterprise",
    "STARTING ARTS WAS FOUNDED IN 2002 WITH THE MISSION TO MAKE ARTS EDUCATION ACCESSIBLE TO ALL STUDENTS IN SAN FRANCISCO BAY AREA SCHOOLS. WE BELIEVE THAT IN ORDER TO STIMULATE AND EDUCATE STUDENTS TO RE": "making arts education accessible across Bay Area schools",
    "SEE STATEMENT 1": None,  # handled by org-specific overrides
    "UStrive democratizes access to social capital so that individuals can achieve their college and career dreams. We do this by connecting mentees to mentors through our virtual mentoring platform and pr": "virtual college mentoring for underserved students",
    "Odyssey School's Mission is: To inspire lasting independent success through accelerated academics and expeditionary learning.": "accelerated academics and expeditionary learning",
    "PROVIDES TRAINING SERVICES TO LOCAL 549 BOILERMAKER APPRENTICE MEMBERS.": "boilermaker apprentice training",
    "Corporation was formed to organize and operate a primary private school, open to the public, based on the Living Wisdom principles of Paramhansa Yogananda (1893-1952) - See Statement 100": "Living Wisdom educational principles for children",
    "TO PROVIDE CHILD CARE SERVICES FOR CHILDREN, AND ALLIED ACTIVITIES OF A SIMILAR NATURE FOR THE CHILDREN, THEIR FAMILIES AND THE COMMUNITY.": "child care services for children and families",
    "The A. Philip Randolph Institute San Francisco (APRISF) is a community-based organization that has served residents of San Francisco, particularly in the Bayview-Hunters Point (BVHP) community since 1": "serving the Bayview-Hunters Point community",
    "CEIDs mission is to maximize the communication potential of young children (0-5 yrs old) who are deaf, hard of hearing, or have severe speech/language delays by providing exemplary early start educati": "early intervention for deaf and hard-of-hearing children",
    "THE SUCCESSFUL OPERATION OF THE CHARITABLE CLINIC IS FUNDAMENTAL TO THE DEVELOPMENT OF FACILITIES, TECHNIQUES, AND TRAINING THAT ARE NECESSARY TO SUPPORT OUR BENEFICIARIES PER OUR MISSION. THROUGH EXP": "sports medicine training and community clinic work",
    "WE PROMISE THAT BY 2028 WE WILL HAVE CLOSED EDUCATIONAL EQUITY GAPS AT EVERY STAGE ALONG A STUDENT'S CRADLE TO CAREER JOURNEY. THAT'S WHY MORE THAN 100 SCHOOL DISTRICTS, COMMUNITY MEMBERS, NONPROFITS,": "closing educational equity gaps from cradle to career in Marin",
    "COALITION OF EMPLOYERS WITH MISSION TO BUILD COLLECTIVE INTELLIGENCE REGARDING CORPORATE AND EMPLOYEE BENEFITS AND TO ADDRESS INTERNATIONAL BENEFIT ISSUES FROM THE PERPECTIVE OF MULTI-NATIONAL EMPLOYE": "employer benefits and workforce strategy in Silicon Valley",
    "A NONPROFIT PUBLIC BENEFIT ORGANIZATION OFFERING PROFESSIONAL EDUCATION IN TRADITIONAL CHINESE MEDICIEN WITHIN THE CONTEXT OF CHINESE CULTURE. IT IS A PROFESSIONAL GRADUATE SCHOOL IN PREPARING GRADUAT": "traditional Chinese medicine education",
    "Child Education Center is a licensed daycare serving children ages 6 weeks thru 6 years.": "daycare for children ages 6 weeks through 6 years",
    "Schools, Mentoring and Resource Team,Inc. champions education equity by supporting students in overcoming systemic barriers on their journey to a college degree.": "championing education equity and college access for students",
    "Our mission is to support K-8 Mandarin instruction within the San Mateo-Foster City School District in order to enable students from all backgrounds the opportunity to develop a mastery of Mandarin Ch": "K-8 Mandarin instruction in San Mateo-Foster City schools",
    "Our mission is to help children, youth, and their families address issues andproblems via Prevention by reaching youth before their problems become crises, andvia Intervention through the provision of": "youth crisis prevention and family intervention in Berkeley",
    "BREAKTHROUGH SILICON VALLEY PREPARES HIGHLY-MOTIVATED STUDENTS TO BECOME FIRST-GENERATION COLLEGE GRADUATES AND INSPIRES EMERGING LEADERS TO BE THE NEXT GENERATION OF EDUCATORS AND ADVOCATES FOR EQUIT": "preparing first-generation college graduates in Silicon Valley",
    "UTILIZING A TRUE PLAY-BASED PHILOSOPHY, WOODROE WOODS SCHOOL PROVIDES A SAFE AND LOVING ENVIRONMENT FOR YOUNG MINDS TO DEVELOP THE COGNITIVE, PHYSICAL, SOCIAL, AND EMOTIONAL SKILLS TO BECOME RESPONSIB": "play-based early childhood development",
    "EDUCATION SUPPORT FOR HOME SCHOOLING": "homeschool education support",
    "AIMC BERKELEY CURRENTLY PROVIDES AN EXTENSIVE CLINICAL PROGRAM, OFFERING MORE THAN 1,000 HOURS OF TRAINING IN ITS CLINICAL AND HERBAL DISPENSARY. OUR COMMITMENT IS TO EDUCATE STUDENTS IN THE SKILLS AN": "integrative medicine education and clinical training",
    "TO ADDRESS BOTH THE ADVANCED INTELLECTUAL ABILITIES AND THE LEARNING CHALLENGES OF EACH STUDENT. STUDENTS CONTINUALLY ENGAGE IN A DYNAMIC, RECIPROCAL LEARNING PROCESS WITH FACULTY.": "education for gifted students with learning challenges",
    "We strengthen groups collaborative media and communications capacity, increase the voiceshare of advocates, experts, and community spokespeople, and provide in-depth media, messaging, and opinion rese": "media strategy and communications for social change",
    "THE DIGITAL EDUCATION PROJECT STRENGTHENS DEMOCRACY BY PREPARING INDIVIDUALS TO MAKE SENSE OF THE PAST AND NAVIGATE OUR DIGITAL PRESENT.": "digital education to strengthen democracy",
    "TO PROVIDE GRADUATE LEVEL EDUCATION IN THE FULL BREADTH OF THE BUDDHIST TRADITION WITH SPECIALIZED EDUCATION SUPPORTING PURE LAND AND CONTEMPORARY SHIN BUDDHIST STUDIES, WHILE ADVANCING JODO SHINSHU B": "graduate-level Buddhist studies",
    "Our mission is to support multicultural, income-diverse families in San Francisco through inclusive, bilingual, play-based early education.": "bilingual, play-based early education for diverse families",
    "Ignite is a movement of women who are ready and eager to become the next generation of political leaders.": "inspiring the next generation of women political leaders",
    "THE MISSION OF THE NAPA VALLEY FARMWORKER FOUNDATION IS TO SUPPORT AND PROMOTE NAPA VALLEY'S VINEYARD WORKERS THROUGH EDUCATION AND PROFESSIONAL DEVELOPMENT.": "education and professional development for vineyard workers",
    "TO PERFORM THE CHARITABLE FUNCTIONS OF AND CARRY OUT THE CHARITABLE PURPOSES OF EAH INC., A RELATED CALIFORNIA PUBLIC BENEFIT CORPORATION.": "affordable housing development",
    "A Home Within is the only national organization dedicated solely to meeting the emotional needs of foster youth.": "meeting the emotional needs of foster youth",
    "THE FOUNDATION'S MISSION IS TO SUPPORT THE NATIONAL UNIVERSITY OF SINGAPORE'S (NUS) MISSION OF GLOBAL EDUCATION, RESEARCH AND SERVICE BY FACILITATING CHARITABLE GIVING AND FOSTERING SUPPORT FROM US-BA": "supporting global education at the National University of Singapore",
    "TO PROVIDE STUDENTS WITH A LEARNING CULTURE WHICH IS GROUNDED IN MONTESSORI PHILOSOPHY. THIS HISTORICALLY PROVEN EDUCATIONAL MODEL SUPPORTS THE WHOLE CHILD, CREATES LIFELONG LEARNERS, AND EDUCATES FOR": "Montessori-grounded whole-child education",
    "THE MISSION OF KID STREET LEARNING CENTER (KSLC) IS TO PROVIDE CHILDREN & FAMILIES LVING IN EXTREME CRISIS A CARING AND SUPPORTIVE EDUCATIONAL COMMUNITY TO CALL HOME WHILE MEETING THE CHILDREN'S EMOTI": "supporting children and families living in extreme crisis",
    "TRANSFORM EAST OAKLAND INTO A HEALTHY AND ECONOMICALLY ROBUST COMMUNITY BY DEVELOPING THE LEADERSHIP OF YOUTH AND YOUNG ADULTS AND IMPROVING THE SYSTEMS THAT IMPACT THEM.": "developing youth leadership in East Oakland",
    "The Scandinavian School in San Francisco provides a place where children and adults can meet to be exposed to Scandinavian languages, traditions and culture. Classes are conducted solely in the Scandi": "Scandinavian language and cultural education for children",
    "Private middle school for boys": "independent middle school education for boys",
    "GPP works within and across social movement sectorsto foster greater alignment around shared goals and strategies for transformational change.Through a combination of strategic consulting and politica": "cross-sector strategy for transformational social change",
    "TO FOCUS ON THE TOTAL DEVELOPMENT OF CHILDREN SOCIALLY, MORALLY, PHYSCIALLY, AND INTECTUALLY.": "holistic child development",
    "EDUCATION": None,  # too generic, handled by org-specific overrides
    "Family Engagement Lab's mission is to help schools ignite the potential of millions of families to support their child's learning.": "helping families support their child's learning",
    "Our Mission Is To Partner With Parents To Educate, Develop, And Nurture Students For A Life Of Christ-centered Excellence In Academics, Christian Character, And Servant Leadership.": "Christ-centered academic excellence and servant leadership",
    "TO EDUCATE SECONDARY STUDENTS WITH NLD, ASPERGER'S SYNDROME AND OTHER NEUROCOGNITIVE DISORDERS IN A PROGRAM THAT EQUALLY EMPHASIZES ACADEMICS, SOCIAL COMPETENCY AND PRAGMATIC LANGUAGE DEVELOPMENT.": "education for students with Asperger's and neurocognitive differences",
    "The Organization's specific and primary mission is to develop and implement programs to encourage students, including stop-outs and dropouts to pursue education and trainings to assure them a chance o": "construction career training for students re-entering education",
    "COASTSIDE CHILDREN'S PROGRAM'S (CCP) MISSION IS TO PROVIDE A SAFE AND CARING ENVIRONMENT WHERE CHILDREN LEARN THROUGH EXPERIENCE, PLAY, AND FRIENDSHIPS - BUILDING A FOUNDATION FOR SUCCESS IN SCHOOL AN": "play-based learning and school readiness on the Coastside",
    "The Burkard School provides a specialized education for children who need extra support with social-emotional learning, self-regulation, and executive functioning.": "specialized education for social-emotional learning and self-regulation",
    "To operate a preschool for children between the ages of 2-5.": "preschool education for children ages 2-5",
    "ASEP's mission is to provide the community with equal access to quality after school enrichment programming. ASEP combines Arts Enrichment, Academic Support and Health & Wellness into a comprehensive": "after-school arts, academics, and wellness programs",
    "THE FOUNDATION PROVIDES SCHOLARSHIPS FOR CCSF STUDENTS, BUILDS AND IMPLEMENTS ENDOWMENT FUNDS TO SUPPORT CCSF STUDENTS AND THE COLLEGE, AND ENLISTS CONTINUING SUPPORT FOR THE COLLEGE WITHIN THE COMMUN": "scholarships and endowments for City College students",
    "THE MISSION OF THE ORGANIZATION IS TO USE THE GOLDEN GATE PARK GOLFCOURSE TO PROVIDE CREATIVE AND INNOVATIVE PROGRAMS TO DISADVANTAGEDSAN FRANCISCO YOUTH AND THE GENERAL COMMUNITY.": "innovative youth programs through golf in Golden Gate Park",
    "A CATHOLIC, COLLEGE PREPARATORY COMMUNITY, EDUCATING STUDENTS IN THE WHOLENESS OF BODY, MIND, AND SOUL, CONSISTENT WITH THE TEACHINGS OF THE CATHOLIC CHURCH, MANIFESTED BY FAITH, LEADERSHIP AND SERVIC": "Catholic college prep education",
    "LOTUS BLOOM IS A FAMILY RESOURCE CENTER WHERE CHILDREN AGED 0-5 AND THEIR PARENTS/CAREGIVERS LEARN, PLAY, GAIN CONFIDENCE AND FIND CONNECTIONS TO OTHER NEW PARENTS, HEALTH RESOURCES, EDUCATION RESOURC": "family resources and early learning for children 0-5",
    "Our mission is to make hands-on food education accessible everywhere children gather to learn so that kids have the knowledge and confidence to make healthy food choices for life. The Charlie Cart Pro": "hands-on food education for children",
    "TO SUPPORT THE DEVELOPMENT, DISTRIBUTION, AND ADOPTION OF OPEN SOURCE SOFTWARE FOR USE IN ROBOTICS RESEARCH, EDUCATION, AND PRODUCT DEVELOPMENT.": "open-source robotics software for education and research",
    "The GNOME Foundation is a non-profit organization that believes in a world where everyone is empowered by technology they can trust. We do this by building a diverse and sustainable free software pers": "free and open-source software for everyone",
    "Preschool education": "preschool education",
    "Islamic Scholarship Fund's mission is to support American Muslim storytelling & increase representation in fields & occupations that make public policy & influence public opinion by providing scholars": "scholarships for American Muslim students in policy and storytelling",
    "TO CREATE ECONOMIC ADVANCEMENT FOR MULTIPLE GENERATIONS, BY PARTNERING WITH YOUNG MOTHERS TO FURTHER THEIR EDUCATION, BUILD LIFE AND CAREER SKILLS, AND NUTURE THEIR CHILDREN'S DEVELOPMENT.": "partnering with young mothers to further their education",
    "To develop children's social, physical, intellectual, emotional, creative, and linguistic skills in preparation for future educational experiences.": "early childhood social and intellectual development",
    "GRANTING SCHOLARSHIPS AND EDUCATIONAL FUNDS TO HELP CHILDREN FROM THE POOREST PARTS OF CHINA AND SOUTHEAST ASIA WITH AN OPPORTUNITY TO COMPLETE A COLLEGE EDUCATION. TO HELP BREAK THE VICIOUS CYCLE OF": "college scholarships for students in rural China and Southeast Asia",
    "BAHIA provides familias with a nurturing, dual language, and learning community where children emerge as engaged contributors to a global society.": "dual-language early learning for Bay Area families",
    "At Compass, social and emotional learning is as important as academic achievement. We teach to students strengths, ensuring academic growth and success with self-advocacy and learning skills, meeting": "social-emotional learning alongside academic achievement",
    "TO PROVIDE EXTENDED DAY CARE FOR CHILDREN WHO ATTEND KINDERGARTEN THROUGH FIFTH GRADE AT FOUR SCHOOLS IN MARTINEZ, CALIFORNIA.": "extended day care for K-5 students in Martinez",
    "Our mission is to partner with educators to help students impacted by poverty demonstrate their full potential on college and career readiness assessments that unlock better opportunities after high s": "college readiness support for students impacted by poverty",
    "PROVIDE QUALITY INFANT CARE AND PRESCHOOL PROGRAMS FOR RESIDENTS OF REDWOODCITY AND ADJACENT AREAS.": "infant care and preschool in Redwood City",
    "As a living tribute to Justice John Paul Stevens, the Foundation works to protect and promote democracy and the rule of law, access to justice, and equality by supporting law students, lawyers, and ot": "promoting democracy, access to justice, and equality",
    "Through lacrosse we support Oakland Public School students in becoming healthy, confident, and self-empowered youth who effectively navigate systems, overcome challenges, and achieve their education a": "using lacrosse to empower Oakland public school students",
    "WE STRIVE TO BUILD CHRISTIAN CHARACTER WHILE PROVIDING A QUALITY EDUCATION, BELIEVING THAT WE ARE TRAINING FOR LIFE. OUR GOAL IS TO TEACH EACH CHILD CHRISTIAN LIVING, SELF-DISCIPLINE, PERSONAL RESPONS": "Christian character education for children",
    "OUR MISSION IS TO SHAPE TECHNOLOGY FOR PUBLIC BENEFIT BY ADVANCING SCIENCES OF CONNECTION AND INTEGRATION.": "shaping technology for public benefit",
    "TO DEVELOP, PROMOTE AND ADMINISTER ELITE SOCCER TRAINING FOR SAN FRANCISCO YOUTH (BOYS AND GIRLS).": "elite soccer training for San Francisco youth",
    "THE MISSION OF EAST PALO ALTO ACADEMY FOUNDATION IS TO ADVOCATE FOR, ADVANCE, AND RAISE FUNDS TO BENEFIT THE STUDENTS AND COLLEGE-ATTENDING GRADUATES OF EAST PALO ALTO ACADEMY, A CHARTER SCHOOL OF THE": "supporting students at East Palo Alto Academy",
    "TO SUPPORT SCHOOL LEADERS TO REACH THEIR HIGHEST POTENTIAL.": "supporting school leaders to reach their potential",
    "EDUCATE PUBLIC ON CONTEMPORARY ISSUES": "public education on contemporary issues",
    "To build a sustainable, geographically distributed dark archive with which to ensure the long-term survival of web-based scholarly publications for the benefit of the greater global research community": "preserving web-based scholarly publications",
    "COLLABORATE WITH CULTURAL INSTITUTIONS AND COMMUNITY ORGANIZATIONS TO CREATE FILMS AND ONLINE MEDIA THAT FOSTER ACTIVE ENGAGEMENT IN CULTURAL AND CIVIC LIFE.": "creating films that foster civic engagement",
    "The Northstar School is a full-time private school that aims to promote human excellence by cultivating students in every grade level who posess a well-trained mind, healthy body, good manners, and ex": "cultivating well-rounded students at every grade level",
    "Khan Schools Network cultivates mastery-based learning through supporting the creation of innovative schools, programs, and curricula, accessible to learners anywhere.": "mastery-based learning and innovative school design",
    "TO PROVIDE LITERARY EVENTS FOR THE BENEFIT OF THE GENERAL PUBLIC AND CHARITABLE ORGANIZATIONS": "literary events for the public",
    "ADVOCATING FOR A GUARANTEED ONE-SEMESTER PERSONAL FINANCE CLASS FOR ALL PUBLIC HIGH SCHOOL STUDENTS TO TAKE BEFORE GRADUATING.": "personal finance education for all high school students",
    "THE ORGANIZATION IS PASSIONATELY COMMITTED TO BUILDING DIVERSE NETWORKS OF LEADERS FOCUSED ON PERSONAL AND COMMUNITY TRANSFORMATION IN ORDER TO CREATE AN INCLUSIVE AND THRIVING SILICON VALLEY.": "building diverse leadership networks in Silicon Valley",
    "The teaching well empowers school systems to more effectively support, retain and leverage the brilliance of their educators. In partnership, we heal adult culture by providing tools for healthy dialo": "empowering school systems to support and retain educators",
    "The mission of Oakland Kids First is to increase youth voice, leadership and power to create engaging and equitable public schools where all students learn and lead.": "increasing youth voice and power in Oakland public schools",
    "THROUGH IN-DEPTH AND THOUGHT-PROVOKING JOURNALISM, PRISM REFLECTS THE LIVED EXPERIENCES OF PEOPLE MOST IMPACTED BY INJUSTICE. AS AN INDEPENDENT AND NONPROFIT NEWSROOM LED BY JOURNALISTS OF COLOR, WE T": "journalism led by journalists of color on injustice",
    "improve the health and sustainability of school communities by greening schoolyards, providing youth environmental education, planting trees, & increasing fresh food access at over 30 public schools .": "greening schoolyards and youth environmental education",
    "Victory Christian Academy's most significant activity is operating a christian school.": "Christian school education",
    "WALDEN CENTER AND SCHOOL IS AN INDEPENDENT , K-6 SCHOOL WHERE DRAMA, MUSIC AND ART SIT ALONGSIDE TRADITIONAL ACADEMIC SUBJECTS AS OUR CORE CURRICULUM. WALDEN'S PHILOSOPHY OF EDUCATION INCORPORATES THE": "K-6 education integrating drama, music, and art",
    "Progressive education in a k-8 school, fostering in-depth learning, inspiring creativity, cultivating critical thinking, social emotional learning, collaboration and communication skills to realize ea": "progressive K-8 education fostering creativity and critical thinking",
    "Mandala offers private-pay, as well as State-Subsidized, childcare for young children in our program. Children are cared for in the nurturing home environment of licensed family childcare homes and th": "nurturing childcare in family home environments",
    "DONUM DEI CLASSICAL ACADEMY'S MISSION IS TO IMPART A RICH CLASSICAL CHRISTIAN CURRICULUM FULL OF SCRIPTURAL TRUTH AND LIFE-GIVING EXPERIENCES.": "classical Christian curriculum education",
    "Our mission is to expose a diverse population of young girls to a high quality STEM academy that inspires confidence in their pursuit of learning throughout the year": "STEM academy programs for young girls",
    "CHILD CARE FOR AT RISK CHILDREN": "child care for at-risk children",
    "TO GRANT SENIOR HIGH SCHOOL SCHOLARSHIPS AND COLLEGE SCHOLARSHIPS TO CHINA RURAL AREA IN POVERTY. ALSO EXTENDED MISSION TO HELP PRESCHOOLERS.": "scholarships for students in rural China",
    "Mission is to increase the number of students finishing high school and enrolling in post secondary education through a combination of academic skills building, and administration of a scholarship pro": "college enrollment support and scholarships for students",
    "To create educational leadership programs for executives in the tech startup sector to create initiatives for increased gender equality in the tech startup sector an d to engage in other charitable an": "leadership programs for women in tech startups",
    "POPS FOUNDATION IS SETUP AS A REFERRAL AND RESOURCE FOR FAMILIES INTERESTED IN ADOPTING ORPHANS FROM CHINA. IT IS OFFERING SERVICES SUCH AS ORGANIZING SEMINARS TO PROVIDE INFORMATION ON THE BASIC REQU": "supporting families adopting orphans from China",
    "FIA was founded to protect access to quality public schools through elevating authentic parent and youth leadership.": "elevating parent and youth leadership in public schools",
    "PROVIDE K-12 GRADE SCHOOL EDUCATION FOR APPROXIMATELY 100 STUDENTS, INCLUDING CLASSROOM INSTRUCTION, CAMPS, FIELD TRIPS, SUMMER CAMPS, AND EXTENDED DAY PROGRAMS.": "K-12 education with hands-on learning and summer camps",
    "TO ENGAGE IN EDUCATIONAL AND CHARITABLE ACTIVITIES BY RAISING FUNDS FOR SETTING UP ENDOWMENTS, CREATING SCHOLARSHIPS, REWARDING TEACHING AND RESEARCH, AND GENERALLY PROMOTING THE DEVELOPMENT OF RESOUR": "scholarships and research endowments for higher education",
    "UP ON TOP offers tuition-free, respectful and inclusive after-school and summer programs to low-income student-learners living primarily in San Francisco's Tenderloin and Western Addition neighborhood": "tuition-free after-school programs in the Tenderloin",
    "TO PROVIDE A SAFE, NURTURING AND ENRICHING ENVIRONMENT FOR K-5 STUDENTS BEFORE AND AFTER SCHOOL AND DURING THE SUMMER.": "before and after-school enrichment for K-5 students",
    "WOOD ROSE ACADEMY'S MISSION IS TO PROVIDE A STRONG ACADEMIC FOUNDATION THROUGH A CLASSICAL CURRICULUM. WE PARTNER WITH PARENTS IN THE INTELLECTUAL, MORAL, PHYSICAL AND SPIRITUAL FORMATION OF THEIR CHI": "classical curriculum K-8 education in partnership with parents",
    "To furnish an academically rigorous, internationally based, year-round day school for K-8 age students.": "rigorous, internationally based K-8 education",
    "IT IS THE MISSION OF NVCF TO PROVIDE RESOURCES TO SUPPORT NAPA VALLEY COLLEGE, ITS PROGRAMS, ITS STUDENTS, AND FACULTY ACTIVITIES IN SUPPORT OF THE MISSION INCLUDE GRANTING SCHOLARSHIPS - CONTINUE ON": "scholarships and student support at Napa Valley College",
    "Through active family involvement in self-governed public education, Manzanita Charter Middle School (A Cooperative Charter), seeks to create a safe, nurturing, and diverse educational community for o": "family-involved cooperative charter middle school education",
    "TO HELP BUILD A BETTER WORLD THROUGH EFFECTIVE EDUCATION.TO GIVE YOUNG PEOPLE A RICH ACADEMIC BACKGROUND, A STRONG SENSE OF ETHICS AND A BROAD RANGE OF ABILITIES TO SUCCESSFULLY LAUNCH THEM INTO HIGHE": "preparing young people with strong academics and ethics",
    "ASSOCIATED STUDENTS, INC. (ASI) OF CALIFORNIA STATE UNIVERSITY, EAST BAY IS THE OFFICIAL VOICE OF THE STUDENTS AT CSU, EAST BAY. ASI REPRESENTS AND ADVOCATES ON BEHALF OF THE INTERESTS, NEEDS, AND CON": "student advocacy at CSU East Bay",
    "Our mission is to increase the capacity of individuals in our community to make things. We provide access to tools, equipment, and knowledge for prototyping, wood and metal fabrications, textiles, ele": "community makerspace access for prototyping and fabrication",
    "We bake and deliver free birthday cakes for underserved children who might not otherwise receive one. In a world where the basics of a positive childhood are often out of reach and youth have little t": "delivering free birthday cakes to underserved children",
    "ROOTED IN THE BAYVIEW HUNTERS POINT NEIGHBORHOOD, RISE UNIVERSITY PREPARATORY OFFERS QUALITY EDUCATION FOR 9TH TO 12TH GRADE.": "quality education for high schoolers in Bayview Hunters Point",
    "Our mission is that BIPOC & LGBTQ transitional age youth 16 -25 will emerge centered, joyful and thriving in the world. We provide opportunity in utilizing art as therapy by celebrating youth empowerm": "art-as-therapy programs for BIPOC and LGBTQ youth",
    "NCWE promotes student success and excellence in education and training by providing the link between policy and practice.": "linking education policy and practice for student success",
    "TO INSPIRE PEOPLE OF ALL AGES AND BACKGROUNDS TO CONNECT WITH THE SEA AND ITS TRIBUTARIES THROUGH EDUCATIONAL PROGRAMS ABROAD TRADITIONAL SAILING VESSELS THAT FOCUS ON THE MARINE SCIENCES, NAUTICAL HE": "marine science education aboard traditional sailing vessels",
    "The Field Semester seeks to reintroduce students to a deeper sense of self and community through an intensive, interdisciplinary education on the land. Academically rigorous and experientially rich, o": "interdisciplinary, land-based education for students",
    "TO PROVIDE A COMPREHENSIVE EARLY CHILDHOOD DEVELOPMENT PROGRAM FOR THE YOUNG CHILDREN OF THE EMPLOYEES OF THE DEPARTMENT OF VETERANS AFFAIRS MEDICAL CENTER.": "early childhood programs for VA employees' children",
    "Silicon Valley Defense Group convenes experts in technology, investment, and government for discussions on how to better collaborate between the various groups. The comments and topics are then publis": "convening tech and defense leaders for collaboration",
    "LEJ promotes ecological health, environmental stewardship, and community development in Southeast San Francisco by creating urban greening, eco-literacy, community stewardship and workforce developmen": "urban greening and eco-literacy in Southeast San Francisco",
    "Private school education and provides academic coaching and classes for private/public/homeschool students.": "academic coaching for private, public, and homeschool students",
    "CHILD EDUCATION & DEVELOPMENT": "child education and development",
    "FOSTER ACTIVITIES TO GENERATE TECHNOLOGIES TO DETER, MITIGATE AND MINIMIZE MASS VIOLENCE IN K-12 SCHOOLS": "school safety technology to prevent mass violence",
    "To support the Fromm Institute For Lifelong Learning program at the University of San Francisco.": "lifelong learning at the University of San Francisco",
    "ARBOR BAY SCHOOL IS A K-8 SCHOOL THAT FOSTERS ACADEMIC AND SOICAL SUCCESS FOR CHILDREN WITH MILD TO MODERATE LEARNING DIFFERENCES, THROUGH MULTISENSORY AND INDIVIDUALIZED INSTURCTION.": "K-8 multisensory education for children with learning differences",
    "THE MONTESSORI SCHOOL OF SILICON VALLEY DEVELOPS THE \"WHOLE CHILD\" BY PROVIDING AN ENRICHED ENVIRONMENT IN WHICH CHILDREN CAN GROW TO FULFILL THEIR INDIVIDUAL POTENTIAL.": "whole-child Montessori education in Silicon Valley",
    "OUR MISSION IS TO PROVIDE HIGH IMPACT, HIGH QUALITY TRAINING AND EDUCATION TO MEMBERS OF THE FIRE SERVICE AT THE LOWEST COST IN ORDER TO REDUCE FIREFIGHTER AND CIVILIAN INJURIES AND DEATHS ON THE EMER": "high-impact firefighter training and education",
    "THE PCF SUPPORTS ACADEMIC EXCELLENCE IN THE PERALTA COMMUNITY COLLEGES DISTRICT BY BUILDING PARTNERSHIPS IN THE REGION TO RAISE FUNDS FOR SCHOLARSHIPS FOR STUDENTS TO THE FOUR DISCTRICT COLLEGES.": "scholarships for Peralta community college students",
    "The mission of the Bay Area Tutoring Association is to develop an academic tutoring workforce for the support of Common Core Math, Literacy, Next Generation Science Standards (NGSS), STEM, Computer Sc": "academic tutoring in math, literacy, and STEM",
    "CATDC creates transformative learning experiences with educators, supporting dynamic collaborations that center equity, spark innovation, and foster impactful teaching and leadership.": "transformative learning experiences for educators",
    "SUPPORTING THE MISSION AND ONGOING WORK OF SOUTH ASIA INSTITUTE OF ADVANCED CHRISTIAN STUDIES, MEETING WITH CHARITABLE FOUNDATINS, TRUSTS, AND THE LIKE TO RAISE FUNDS FOR THE EXPANSION OF SOUTH ASIA I": "supporting Christian higher education in South Asia",
    "to provide and promote the educational, physical, emotional, social and cognitive development of each child.": "holistic early childhood education and development",
    "THE ORGANIZATION STRIVES TO CREATE A NURTURING FAMILY ENVIRONMENT AND GUIDE THE EMOTIONAL AND INTELLECTUAL GROWTH OF YOUNG CHILDREN WITH A FOCUS ON DEVELOPING SELF-CONFIDENCE AND POSITIVE SOCIAL SKILL": "nurturing emotional and intellectual growth in young children",
    "The specific and primary purpose are to engage in the conduct and operation of a school or institution of charitable or religious education and instruction": "charitable education for children",
    "Provide child care and education to low income families.": "child care and education for low-income families",
    "MINDFUL SCHOOL'S MISSION IS TO EMPOWER EDUCATORS TO SPARK CHANGE FROM THE INSIDE OUT BY CULTIVATING AWARENESS, RESILIENCE, AND COMPASSIONATE ACTION.": "empowering educators through mindfulness and resilience",
    "TO PROVIDE TAMIL LANGUAGE EDUCATION TO STUDENTS.": "Tamil language education for students",
    "THIS IS A PRIVATE K-8 CHRISTIAN SCHOOL": "K-8 Christian school education",
    "THE DAY SCHOOL MANAGES FUNDS RAISED BY ST. MATTHEW'S EPISCOPAL DAY SCHOOL IN SUPPORT OF THE SCHOOL'S OPERATIONS, PROGRAMS, AND FACILITIES.": "supporting St. Matthew's Episcopal Day School",
    "The Lowell Alumni Association (LAA) maintains and enhances relationships that serve and support our alumni, students, school and the Lowell community.1.Scholarships: The LAA awards a number of scholar": "alumni scholarships and community at Lowell High School",
    "PROVIDE FINANCIAL AID AND SUPPORT BY WAY OF GIFTS, GRANTS, OR LOANS FOR SELECTED PROGRAMS TO BE PROVIDED BY SCHOOL DISTRICTS OR OTHER PUBLIC OR PRIVATE ORGANIZATIONS OR GROUPS WHICH FURTHER EDUCATIONA": "grants and support for educational programs in Corte Madera-Larkspur",
}

# Org-specific overrides for vague/missing missions
ORG_OVERRIDES = {
    "CREATIVE COMMONS CORPORATION": "open educational resources and licensing",
    "ERUDITE TECHNOLOGY GROUP": "education technology",
    "SEMI FOUNDATION": "semiconductor industry education and workforce development",
    "CONTRA COSTA JEWISH DAY SCHOOL": "Jewish day school education",
    "GOLDEN BRIDGES SCHOOL": "Waldorf-inspired education",
    "ARISE EDUCATIONAL CENTER": "educational programs for the community",
    "TAMALPAIS - STRAWBERRY": "education programs in Marin County",
}


def clean_org_name(name):
    """Title-case org names that are ALL CAPS, preserve mixed case."""
    if not name:
        return name
    name = name.strip()
    # Remove trailing & or incomplete words
    name = re.sub(r'\s+&\s*$', '', name)
    # If all caps, title-case it
    if name == name.upper() and len(name) > 3:
        # Title case with smart handling
        words = name.split()
        result = []
        small_words = {'OF', 'THE', 'AND', 'FOR', 'IN', 'ON', 'AT', 'TO', 'A', 'AN', 'INC', 'DBA'}
        for i, w in enumerate(words):
            if i == 0:
                result.append(w.title())
            elif w in small_words and i != 0:
                if w == 'INC':
                    result.append('Inc.')
                else:
                    result.append(w.lower())
                    # But capitalize if after a dash or if it's the last word
            else:
                result.append(w.title())
        name = ' '.join(result)
    # Remove trailing "Inc" or "Inc." if it makes the name too long
    # Actually, keep it for now
    return name


def get_first_name(full_name):
    """Extract first name from full name."""
    if not full_name:
        return ""
    return full_name.strip().split()[0]


def generate_message(contact_name, org_name, mission):
    """Generate LinkedIn connection message."""
    first_name = get_first_name(contact_name)
    clean_org = clean_org_name(org_name)

    # Get phrase from mission mapping
    mission_key = mission.strip() if mission else ""
    phrase = MISSION_PHRASES.get(mission_key)

    # If no phrase (None = vague mission), try org override
    if phrase is None:
        org_upper = org_name.strip().upper() if org_name else ""
        for org_key, org_phrase in ORG_OVERRIDES.items():
            if org_key in org_upper:
                phrase = org_phrase
                break

    # Final fallback
    if phrase is None:
        phrase = "education and community impact"

    msg = f"Hi {first_name}, I saw {clean_org} is doing great work in {phrase}. Would love to connect."

    # Truncate if over 300 chars
    if len(msg) > 300:
        # Shorten the phrase
        available = 300 - len(f"Hi {first_name}, I saw {clean_org} is doing great work in . Would love to connect.")
        if available > 20:
            phrase = phrase[:available-3] + "..."
            msg = f"Hi {first_name}, I saw {clean_org} is doing great work in {phrase}. Would love to connect."
        else:
            # Org name too long, shorten it
            max_org = 300 - len(f"Hi {first_name}, I saw  is doing great work in {phrase}. Would love to connect.")
            if max_org > 10:
                clean_org = clean_org[:max_org-3] + "..."
            msg = f"Hi {first_name}, I saw {clean_org} is doing great work in {phrase}. Would love to connect."
            if len(msg) > 300:
                msg = msg[:297] + "..."

    return msg


def main():
    path = 'Enhancements/2026-02-23/DATA_2026-02-23_linkedin_prospects_ca_youth_education.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    count = 0
    errors = []
    over_300 = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
        contact = row[1].value   # B = Contact Name
        org = row[5].value       # F = Org Name
        mission = row[6].value   # G = Mission
        existing = row[8].value  # I = My Message

        if not contact:
            continue
        if existing:
            continue  # Don't overwrite existing messages
        if not mission:
            errors.append(f"Row {row_idx}: No mission for {contact} at {org}")
            continue

        msg = generate_message(contact, org, mission)

        if len(msg) > 300:
            over_300 += 1
            errors.append(f"Row {row_idx}: Message {len(msg)} chars for {contact}")

        # Write to column I
        ws.cell(row=row_idx, column=9, value=msg)
        count += 1

    wb.save(path)
    print(f"Generated {count} messages")
    print(f"Over 300 chars: {over_300}")
    if errors:
        print(f"\nWarnings ({len(errors)}):")
        for e in errors[:20]:
            print(f"  {e}")

    # Verify a few samples
    print("\n--- Sample messages ---")
    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2.active
    samples = [3, 5, 10, 50, 100, 200, 300, 400, 500]
    for r in samples:
        if r <= ws2.max_row:
            name = ws2.cell(row=r, column=2).value
            msg = ws2.cell(row=r, column=9).value
            if name and msg:
                print(f"  Row {r} ({name}): [{len(msg)} chars] {msg}")


if __name__ == '__main__':
    main()
