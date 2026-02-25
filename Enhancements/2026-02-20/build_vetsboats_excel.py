#!/usr/bin/env python3
"""
VetsBoats Foundation Shortlist — Excel with priority ranking
Output: DATA_2026-02-20_vetsboats_ranked_prospects.xlsx
"""

import sys
sys.path.insert(0, '/Users/aleckleinman/Documents/TheGrantScout/0. Tools')

import psycopg2
import psycopg2.extras
from xlsx_utils import create_workbook

OUTPUT_DIR = '/Users/aleckleinman/Documents/TheGrantScout/Enhancements/2026-02-20'
CLIENT_EIN = '464194065'

KNOWN_FUNDERS = {
    '680400024', '953059828', '956054953', '841849498', '843805975'
}

VETERAN_KEYWORDS_REGEX = r'(veteran|military|armed forces|service member|wounded warrior|adaptive sailing|therapeutic sailing|adaptive recreation)'

conn = psycopg2.connect(host='localhost', port=5432, database='thegrantscout', user='postgres')
conn.set_session(readonly=True)
cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ---- Build candidate pool ----
cur.execute("""
    SELECT DISTINCT foundation_ein FROM f990_2025.calc_client_foundation_scores
    WHERE client_ein = %s
""", (CLIENT_EIN,))
source_a = {r['foundation_ein'] for r in cur.fetchall()}

cur.execute("""
    SELECT DISTINCT foundation_ein FROM f990_2025.fact_grants
    WHERE LOWER(purpose_text) ~ %s
""", (VETERAN_KEYWORDS_REGEX,))
source_b = {r['foundation_ein'] for r in cur.fetchall()}

all_eins = list(source_a | source_b)
source_map = {}
for e in all_eins:
    in_a, in_b = e in source_a, e in source_b
    source_map[e] = 'both' if in_a and in_b else ('sibling' if in_a else 'keyword')

# ---- Latest filing data ----
cur.execute("""
    SELECT DISTINCT ON (ein)
        ein, business_name, state, website_url,
        total_assets_eoy_amt, fmv_assets_eoy_amt,
        only_contri_to_preselected_ind, tax_year,
        mission_desc, activity_or_mission_desc,
        phone_num, app_contact_email, app_contact_name,
        app_submission_deadlines, app_restrictions, app_form_requirements
    FROM f990_2025.pf_returns
    WHERE ein = ANY(%s)
    ORDER BY ein, tax_year DESC, return_timestamp DESC NULLS LAST
""", (all_eins,))
filings = {r['ein']: dict(r) for r in cur.fetchall()}

# ---- Filter 1: Open to applicants ----
f1_pass = {e for e, d in filings.items() if d.get('only_contri_to_preselected_ind') is not True}

# ---- Filter 2: CA geography ----
cur.execute("""
    SELECT foundation_ein, COUNT(*) as cnt
    FROM f990_2025.fact_grants WHERE foundation_ein = ANY(%s) AND recipient_state = 'CA'
    GROUP BY foundation_ein
""", (all_eins,))
ca_grants = {r['foundation_ein']: r['cnt'] for r in cur.fetchall()}
f2_pass = {e for e in filings if ca_grants.get(e, 0) > 0 or filings[e].get('state') == 'CA'}

# ---- Filter 3: PF-to-PF grants ----
cur.execute("""
    SELECT fg.foundation_ein,
           COUNT(DISTINCT fg.id) as pf_cnt,
           COALESCE(SUM(fg.amount), 0) as pf_total
    FROM f990_2025.fact_grants fg
    WHERE fg.foundation_ein = ANY(%s) AND fg.recipient_ein IS NOT NULL
      AND EXISTS (SELECT 1 FROM f990_2025.pf_returns pr WHERE pr.ein = fg.recipient_ein)
    GROUP BY fg.foundation_ein
""", (all_eins,))
pf_grants = {r['foundation_ein']: {'cnt': r['pf_cnt'], 'total': int(r['pf_total'])} for r in cur.fetchall()}
f3_pass = set(pf_grants.keys())

passed_all = f1_pass & f2_pass & f3_pass
near_miss = {e for e in filings if sum([e in f1_pass, e in f2_pass, e in f3_pass]) == 2}

# ---- Additional metrics ----
# Veteran grants
cur.execute("""
    SELECT foundation_ein, COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total
    FROM f990_2025.fact_grants WHERE foundation_ein = ANY(%s)
      AND LOWER(purpose_text) ~ %s
    GROUP BY foundation_ein
""", (all_eins, VETERAN_KEYWORDS_REGEX))
vet = {r['foundation_ein']: {'cnt': r['cnt'], 'total': int(r['total'])} for r in cur.fetchall()}

# Total grants
cur.execute("""
    SELECT foundation_ein, COUNT(*) as cnt FROM f990_2025.fact_grants
    WHERE foundation_ein = ANY(%s) GROUP BY foundation_ein
""", (all_eins,))
total_grants = {r['foundation_ein']: r['cnt'] for r in cur.fetchall()}

# Profiles
cur.execute("""
    SELECT ein, total_giving_5yr, unique_recipients_5yr, openness_score,
           new_recipients_5yr, last_active_year, giving_trend, accepts_applications,
           median_grant, repeat_rate
    FROM f990_2025.calc_foundation_profiles WHERE ein = ANY(%s)
""", (all_eins,))
profiles = {r['ein']: dict(r) for r in cur.fetchall()}

# Sibling data
cur.execute("""
    SELECT foundation_ein, target_grants_to_siblings, lasso_score
    FROM f990_2025.calc_client_foundation_scores
    WHERE client_ein = %s AND foundation_ein = ANY(%s)
""", (CLIENT_EIN, all_eins))
siblings = {r['foundation_ein']: dict(r) for r in cur.fetchall()}

cur.close()
conn.close()

# ---- Assemble rows for passed + near-miss ----
def clean_url(url):
    if not url or url.lower() in ('n/a', 'none', '0', 'null', ''):
        return ''
    return url

def clean_text(t):
    if not t or t.lower() in ('n/a', 'none', ''):
        return ''
    return t

rows = []
for ein in list(passed_all) + list(near_miss):
    f = filings.get(ein, {})
    p = profiles.get(ein, {})
    s = siblings.get(ein, {})

    # Determine section
    passes = sum([ein in f1_pass, ein in f2_pass, ein in f3_pass])
    section = 'PASSED ALL 3' if passes == 3 else 'NEAR MISS (2/3)'

    # Failed filter
    failed = []
    if ein not in f1_pass: failed.append('Preselected Only')
    if ein not in f2_pass: failed.append('No CA Grants')
    if ein not in f3_pass: failed.append('No PF-to-PF')
    failed_str = '; '.join(failed) if failed else ''

    # Mission description (longer of two)
    m1 = f.get('mission_desc') or ''
    m2 = f.get('activity_or_mission_desc') or ''
    mission = m1 if len(m1) >= len(m2) else m2

    assets = f.get('total_assets_eoy_amt') or f.get('fmv_assets_eoy_amt')
    annual_giving = int(p.get('total_giving_5yr', 0) or 0) // 5 if p.get('total_giving_5yr') else None

    # Priority score (higher = review first)
    # Weights: veteran $ (30%), giving capacity (25%), CA focus (15%), PF confidence (15%), contactability (15%)
    import math

    score = 0
    vet_cnt = vet.get(ein, {}).get('cnt', 0)
    vet_tot = vet.get(ein, {}).get('total', 0)
    has_email = bool(clean_text(f.get('app_contact_email', '')))
    has_website = bool(clean_url(f.get('website_url', '')))
    new_recip = p.get('new_recipients_5yr') or 0
    last_yr = p.get('last_active_year') or 0
    accepts = p.get('accepts_applications')
    trend = p.get('giving_trend', '')
    tot_grants = total_grants.get(ein, 0)

    # --- Veteran mission (0-30 pts): log-scaled amount, not just count ---
    if vet_tot > 0:
        # $1K=3pts, $10K=6, $100K=10, $1M=20, $5M+=30
        score += min(30, int(math.log10(max(vet_tot, 1)) * 6))

    # --- Giving capacity (0-25 pts): penalize tiny, reward meaningful ---
    ag = annual_giving or 0
    if ag >= 10_000_000: score += 25      # $10M+/yr
    elif ag >= 1_000_000: score += 20     # $1M+
    elif ag >= 500_000: score += 15       # $500K+
    elif ag >= 100_000: score += 10       # $100K+
    elif ag >= 50_000: score += 5         # $50K+
    # <$50K/yr gets 0 (too small to realistically fund VetsBoats)

    # --- CA geographic focus (0-15 pts): CA grants as % of total ---
    ca_cnt = ca_grants.get(ein, 0)
    if tot_grants > 0:
        ca_pct = ca_cnt / tot_grants
        if ca_pct >= 0.30: score += 15     # 30%+ grants in CA
        elif ca_pct >= 0.10: score += 10   # 10%+
        elif ca_pct >= 0.03: score += 5    # 3%+
        elif ca_cnt >= 10: score += 3      # at least some CA presence

    # --- PF-to-PF confidence (0-15 pts): more PF grants = more trustworthy ---
    pf_cnt = pf_grants.get(ein, {}).get('cnt', 0)
    if pf_cnt >= 20: score += 15
    elif pf_cnt >= 10: score += 12
    elif pf_cnt >= 5: score += 9
    elif pf_cnt >= 3: score += 6
    elif pf_cnt >= 1: score += 3

    # --- Contactability & actionability (0-15 pts) ---
    if has_email: score += 6
    if has_website: score += 3
    if accepts is True: score += 3
    if last_yr and last_yr >= 2023: score += 3

    # --- Bonuses ---
    if new_recip >= 10: score += 2         # actively taking new grantees
    if trend == 'growing': score += 2      # growing funder
    if source_map.get(ein) in ('both', 'sibling'): score += 2  # sibling match

    # --- Penalties ---
    if ein in KNOWN_FUNDERS: score -= 200
    if ag and ag < 10_000: score -= 10     # micro-foundation penalty

    rows.append({
        'priority_score': score,
        'section': section,
        'foundation_name': f.get('business_name', ''),
        'ein': ein,
        'state': f.get('state', ''),
        'website': clean_url(f.get('website_url', '')),
        'assets': int(assets) if assets else None,
        'annual_giving': annual_giving,
        'median_grant': int(p['median_grant']) if p.get('median_grant') else None,
        'vet_grants': vet_cnt,
        'vet_amount': vet_tot,
        'pf_grants': pf_grants.get(ein, {}).get('cnt', 0),
        'pf_amount': pf_grants.get(ein, {}).get('total', 0),
        'ca_grants': ca_grants.get(ein, 0),
        'total_grants': total_grants.get(ein, 0),
        'new_recipients_5yr': new_recip,
        'last_active_year': last_yr or None,
        'openness_score': float(p['openness_score']) if p.get('openness_score') else None,
        'accepts_apps': 'Yes' if accepts is True else ('No' if accepts is False else ''),
        'giving_trend': trend or '',
        'contact_email': clean_text(f.get('app_contact_email', '')),
        'contact_name': clean_text(f.get('app_contact_name', '')),
        'phone': clean_text(f.get('phone_num', '')),
        'app_deadlines': clean_text(f.get('app_submission_deadlines', '')),
        'app_restrictions': clean_text(f.get('app_restrictions', '')),
        'sibling_grants': s.get('target_grants_to_siblings', 0) or 0,
        'source': source_map.get(ein, ''),
        'known_funder': 'YES' if ein in KNOWN_FUNDERS else '',
        'failed_filter': failed_str,
        'mission': mission[:300] if mission else '',
        'latest_tax_year': f.get('tax_year'),
    })

# Sort: section (PASSED first), then priority_score DESC
rows.sort(key=lambda r: (0 if r['section'] == 'PASSED ALL 3' else 1, -r['priority_score'], -r['vet_grants'], -r['pf_grants']))

# ---- Build Excel ----
headers = [
    'Priority', 'Section', 'Foundation Name', 'EIN', 'State', 'Website',
    'Total Assets', 'Annual Giving', 'Median Grant',
    'Veteran Grants', 'Veteran $', 'PF Grants', 'PF $',
    'CA Grants', 'Total Grants', 'New Recipients (5yr)',
    'Last Active Year', 'Openness', 'Accepts Apps', 'Trend',
    'Contact Email', 'Contact Name', 'Phone',
    'App Deadlines', 'App Restrictions',
    'Sibling Grants', 'Source', 'Known Funder', 'Failed Filter',
    'Mission/Activity', 'Tax Year'
]

data = []
for r in rows:
    data.append([
        r['priority_score'], r['section'], r['foundation_name'], r['ein'], r['state'], r['website'],
        r['assets'], r['annual_giving'], r['median_grant'],
        r['vet_grants'], r['vet_amount'], r['pf_grants'], r['pf_amount'],
        r['ca_grants'], r['total_grants'], r['new_recipients_5yr'],
        r['last_active_year'], r['openness_score'], r['accepts_apps'], r['giving_trend'],
        r['contact_email'], r['contact_name'], r['phone'],
        r['app_deadlines'], r['app_restrictions'],
        r['sibling_grants'], r['source'], r['known_funder'], r['failed_filter'],
        r['mission'], r['latest_tax_year'],
    ])

col_formats = {
    'Total Assets': 'currency',
    'Annual Giving': 'currency',
    'Median Grant': 'currency',
    'Veteran $': 'currency',
    'PF $': 'currency',
    'Veteran Grants': 'number',
    'PF Grants': 'number',
    'CA Grants': 'number',
    'Total Grants': 'number',
    'New Recipients (5yr)': 'number',
    'Sibling Grants': 'number',
    'Priority': 'number',
}

passed_count = sum(1 for r in rows if r['section'] == 'PASSED ALL 3')
near_count = sum(1 for r in rows if r['section'] != 'PASSED ALL 3')

output_path = f"{OUTPUT_DIR}/DATA_2026-02-20_vetsboats_ranked_prospects.xlsx"

# ---- Build workbook manually (no table object) to avoid hyperlink corruption ----
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "VetsBoats Prospects"
ws.sheet_view.zoomScale = 140

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name='Calibri', size=10, color="2C3E50")
LINK_FONT = Font(name='Calibri', size=10, color='2980B9', underline='single')
NO_WRAP = Alignment(wrap_text=False, vertical="center")
THIN = Side(border_style="thin", color="000000")
MED = Side(border_style="medium", color="000000")

FMT = {
    'Total Assets': '"$"#,##0', 'Annual Giving': '"$"#,##0', 'Median Grant': '"$"#,##0',
    'Veteran $': '"$"#,##0', 'PF $': '"$"#,##0',
    'Veteran Grants': '#,##0', 'PF Grants': '#,##0', 'CA Grants': '#,##0',
    'Total Grants': '#,##0', 'New Recipients (5yr)': '#,##0', 'Sibling Grants': '#,##0',
    'Priority': '#,##0',
}

website_idx = headers.index('Website')
email_idx = headers.index('Contact Email')
start_row, start_col = 2, 2

# Write headers
for ci, h in enumerate(headers):
    c = ws.cell(row=start_row, column=start_col + ci, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = NO_WRAP

# Write data
for ri, row_data in enumerate(data):
    for ci, val in enumerate(row_data):
        c = ws.cell(row=start_row + 1 + ri, column=start_col + ci, value=val)
        c.alignment = NO_WRAP
        fmt = FMT.get(headers[ci])
        if fmt:
            c.number_format = fmt

        # Hyperlinks
        if ci == website_idx and val and str(val).strip():
            href = val if str(val).startswith('http') else f'http://{val}'
            c.hyperlink = href
            c.font = LINK_FONT
        elif ci == email_idx and val and '@' in str(val):
            c.hyperlink = f'mailto:{val}'
            c.font = LINK_FONT
        else:
            c.font = DATA_FONT

# Borders
end_row = start_row + len(data)
end_col = start_col + len(headers) - 1
for r in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        c = ws.cell(row=r, column=col)
        t = MED if r == start_row else THIN
        b = MED if r in (start_row, end_row) else THIN
        l = MED if col == start_col else THIN
        ri = MED if col == end_col else THIN
        c.border = Border(top=t, bottom=b, left=l, right=ri)

# Column widths
for ci in range(len(headers)):
    cl = get_column_letter(start_col + ci)
    max_len = len(headers[ci])
    for rd in data[:50]:  # sample first 50 rows
        if ci < len(rd) and rd[ci] is not None:
            max_len = max(max_len, len(str(rd[ci])))
    ws.column_dimensions[cl].width = min(max_len + 4, 50)
ws.column_dimensions['A'].width = 2

# Autofilter (simple filter row, no Table object)
ws.auto_filter.ref = f"B2:{get_column_letter(end_col)}{end_row}"

# Freeze panes
ws.freeze_panes = ws.cell(row=3, column=2)

wb.save(output_path)
print(f"Created: {output_path}")

# ---- Print top 10 for quick review ----
print(f"\nPassed all 3: {passed_count}")
print(f"Near misses:  {near_count}")
print(f"Total rows:   {len(rows)}")

top10 = [r for r in rows if r['section'] == 'PASSED ALL 3'][:10]
print(f"\n{'='*80}")
print(f"TOP 10 — Start here:")
print(f"{'='*80}")
for i, r in enumerate(top10, 1):
    url = r['website'] or '(no website)'
    email = r['contact_email'] or '(no email)'
    print(f"\n{i}. {r['foundation_name']}")
    print(f"   Score: {r['priority_score']} | EIN: {r['ein']} | {r['state']}")
    print(f"   Vet Grants: {r['vet_grants']} (${r['vet_amount']:,}) | PF Grants: {r['pf_grants']} | CA: {r['ca_grants']}")
    print(f"   Assets: ${r['assets']:,}" if r['assets'] else "   Assets: N/A", end="")
    print(f" | Annual Giving: ${r['annual_giving']:,}" if r['annual_giving'] else " | Annual Giving: N/A")
    print(f"   Accepts Apps: {r['accepts_apps']} | Active: {r['last_active_year']} | Trend: {r['giving_trend']}")
    print(f"   Website: {url}")
    print(f"   Email: {email}")
