#!/usr/bin/env python3
"""Import outreach history from Calls-2.xlsx into campaign_prospect_status."""

import openpyxl
import psycopg2
from datetime import datetime

XLSX = "/Users/aleckleinman/Documents/TheGrantScout/Calls-2.xlsx"

conn = psycopg2.connect(
    host='localhost', port=5432, database='thegrantscout',
    user='postgres', password='postgres'
)
conn.autocommit = False
cur = conn.cursor()
cur.execute("SET search_path TO f990_2025")

wb = openpyxl.load_workbook(XLSX, data_only=True)
stats = {"calls": 0, "sheet2": 0, "sheet3_enriched": 0, "sheet4": 0, "sheet5": 0,
         "skipped_dup": 0, "ein_matched": 0}

OUTCOME_MAP = {
    "1. Not Contacted": "queued",
    "2. VM": "contacted",
    "3. Talked to Someone": "contacted",
    "4. Reached Decision Maker": "contacted",
    "5. Interested": "replied",
}


def clean_ein(val):
    if val is None:
        return None
    s = str(val).strip().replace("-", "").replace(" ", "")
    if not s or s.lower() in ("none", "n/a", "0"):
        return None
    # Remove .0 from float conversion
    if "." in s:
        s = s.split(".")[0]
    # EINs should be 9 digits, pad with leading zeros
    if s.isdigit() and len(s) < 9:
        s = s.zfill(9)
    return s if s.isdigit() else None


def clean_email(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or "@" not in s or s.lower() in ("none", "n/a"):
        return None
    return s.lower()


def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("none", "n/a"):
        return None
    return s


def clean_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def exists_by_email_vertical(email, vertical):
    if email is None:
        return False
    cur.execute(
        "SELECT 1 FROM campaign_prospect_status WHERE email = %s AND vertical = %s LIMIT 1",
        (email, vertical)
    )
    return cur.fetchone() is not None


def exists_by_ein_vertical(ein, vertical):
    if ein is None:
        return False
    cur.execute(
        "SELECT 1 FROM campaign_prospect_status WHERE ein = %s AND vertical = %s LIMIT 1",
        (ein, vertical)
    )
    return cur.fetchone() is not None


def insert_cps(ein, email, org_name, org_type, vertical, campaign_status,
               initial_sent_at, notes, source_file):
    # Truncate org_type to 20 chars (VARCHAR(20) limit)
    if org_type and len(org_type) > 20:
        org_type = org_type[:20]
    cur.execute("""
        INSERT INTO campaign_prospect_status
            (ein, email, org_name, org_type, vertical, campaign_status,
             initial_sent_at, notes, source_file)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (ein, email, org_name, org_type, vertical, campaign_status,
          initial_sent_at, notes, source_file))


# ─── SHEET: Calls (main — rows 3-168, header in row 2) ───
print("=== Importing Calls sheet ===")
ws = wb["Calls"]
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    org_name = clean_str(row[1].value)  # B
    if not org_name:
        continue

    email = clean_email(row[7].value)   # H
    ein = clean_ein(row[11].value)      # L
    org_type = clean_str(row[17].value) # R
    outcome = clean_str(row[18].value)  # S
    date_contacted = clean_date(row[3].value)  # D
    call_notes = clean_str(row[19].value)      # T

    campaign_status = OUTCOME_MAP.get(outcome, "queued") if outcome else "queued"

    # Dedup: skip if already exists by email or EIN
    if exists_by_email_vertical(email, "cold_call"):
        stats["skipped_dup"] += 1
        continue
    if exists_by_ein_vertical(ein, "cold_call"):
        stats["skipped_dup"] += 1
        continue

    insert_cps(ein, email, org_name, org_type, "cold_call", campaign_status,
               date_contacted, call_notes, "Calls-2.xlsx/Calls")
    stats["calls"] += 1

print(f"  Calls: {stats['calls']} imported, {stats['skipped_dup']} skipped (dup)")


# ─── SHEET: Sheet2 (17 researched foundations, header in row 2) ───
print("\n=== Importing Sheet2 ===")
ws2 = wb["Sheet2"]
for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row):
    org_name = clean_str(row[1].value)  # B
    if not org_name:
        continue

    ein = clean_ein(row[8].value)       # I
    email = clean_email(row[15].value)  # P
    date_contacted = clean_date(row[3].value)  # D

    # Skip if already imported from Calls
    if exists_by_email_vertical(email, "cold_call"):
        stats["skipped_dup"] += 1
        continue
    if exists_by_ein_vertical(ein, "cold_call"):
        stats["skipped_dup"] += 1
        continue

    insert_cps(ein, email, org_name, None, "cold_call", "queued",
               date_contacted, None, "Calls-2.xlsx/Sheet2")
    stats["sheet2"] += 1

print(f"  Sheet2: {stats['sheet2']} imported")


# ─── SHEET: Sheet3 (84 contact records — enrich fp2, do NOT create CPS rows) ───
print("\n=== Processing Sheet3 (contact enrichment) ===")
ws3 = wb["Sheet3"]
for row in ws3.iter_rows(min_row=3, max_row=ws3.max_row):
    org_name = clean_str(row[1].value)  # B
    if not org_name:
        continue

    ein = clean_ein(row[9].value)       # J
    contact_email = clean_email(row[6].value)  # G

    if ein and contact_email:
        cur.execute("""
            UPDATE foundation_prospects2
            SET app_contact_email = %s
            WHERE ein = %s AND (app_contact_email IS NULL OR app_contact_email = '')
        """, (contact_email, ein))
        if cur.rowcount > 0:
            stats["sheet3_enriched"] += 1

print(f"  Sheet3: {stats['sheet3_enriched']} fp2 records enriched with contact email")


# ─── SHEET: Sheet4 (106 contacts, no email, org in C, EIN in K) ───
print("\n=== Importing Sheet4 ===")
ws4 = wb["Sheet4"]
for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
    org_name = clean_str(row[2].value)  # C
    if not org_name:
        continue

    ein = clean_ein(row[10].value)      # K
    state = clean_str(row[8].value)     # I
    org_type = clean_str(row[14].value) # O

    # No email column in Sheet4
    if exists_by_ein_vertical(ein, "cold_call"):
        stats["skipped_dup"] += 1
        continue

    # For rows without EIN, check by org name (approximate)
    if ein is None:
        cur.execute(
            "SELECT 1 FROM campaign_prospect_status WHERE org_name = %s AND vertical = 'cold_call' LIMIT 1",
            (org_name,)
        )
        if cur.fetchone():
            stats["skipped_dup"] += 1
            continue

    insert_cps(ein, None, org_name, org_type, "cold_call", "queued",
               None, None, "Calls-2.xlsx/Sheet4")
    stats["sheet4"] += 1

print(f"  Sheet4: {stats['sheet4']} imported")


# ─── SHEET: Sheet5 (50 priority foundations, header row 3, data starts row 4) ───
print("\n=== Importing Sheet5 ===")
ws5 = wb["Sheet5"]
for row in ws5.iter_rows(min_row=4, max_row=ws5.max_row):
    org_name = clean_str(row[2].value)  # C
    if not org_name:
        continue

    ein = clean_ein(row[15].value)      # P
    email = clean_email(row[9].value)   # J
    state = clean_str(row[3].value)     # D

    if exists_by_email_vertical(email, "cold_call"):
        stats["skipped_dup"] += 1
        continue
    if exists_by_ein_vertical(ein, "cold_call"):
        stats["skipped_dup"] += 1
        continue

    insert_cps(ein, email, org_name, None, "cold_call", "queued",
               None, None, "Calls-2.xlsx/Sheet5")
    stats["sheet5"] += 1

print(f"  Sheet5: {stats['sheet5']} imported")

conn.commit()

# ─── Final Report ───
print("\n" + "=" * 60)
print("IMPORT SUMMARY")
print("=" * 60)
total_added = stats["calls"] + stats["sheet2"] + stats["sheet4"] + stats["sheet5"]
print(f"  Calls sheet:  {stats['calls']} rows imported")
print(f"  Sheet2:       {stats['sheet2']} rows imported")
print(f"  Sheet3:       {stats['sheet3_enriched']} fp2 emails enriched (no CPS rows)")
print(f"  Sheet4:       {stats['sheet4']} rows imported")
print(f"  Sheet5:       {stats['sheet5']} rows imported")
print(f"  Skipped:      {stats['skipped_dup']} duplicates")
print(f"  TOTAL ADDED:  {total_added}")

# Count totals
cur.execute("""
    SELECT COUNT(DISTINCT ein) as unique_eins,
           COUNT(DISTINCT email) as unique_emails,
           COUNT(*) as total_rows
    FROM campaign_prospect_status
    WHERE vertical = 'cold_call'
""")
r = cur.fetchone()
print(f"\n  Cold-call records: {r[2]} total")
print(f"  Unique EINs in suppress: {r[0]}")
print(f"  Unique emails in suppress: {r[1]}")

# Also show overall suppress list
cur.execute("""
    SELECT COUNT(DISTINCT ein) as unique_eins,
           COUNT(DISTINCT email) as unique_emails,
           COUNT(*) as total_rows
    FROM campaign_prospect_status
""")
r = cur.fetchone()
print(f"\n  ALL records: {r[2]} total")
print(f"  ALL unique EINs: {r[0]}")
print(f"  ALL unique emails: {r[1]}")

cur.close()
conn.close()
print("\nDone.")
